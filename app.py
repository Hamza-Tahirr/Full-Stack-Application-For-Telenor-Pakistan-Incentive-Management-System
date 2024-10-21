from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, Response, stream_with_context
from hdbcli import dbapi
import math
import pandas as pd
from datetime import datetime
from jinja2 import Environment, PackageLoader, select_autoescape
from hdbcli.dbapi import IntegrityError, Error
from io import BytesIO
import xlsxwriter
# from werkzeug.security import generate_password_hash, check_password_hash
import os
import time
import io
import hashlib
import logging
from hdbcli import dbapi
from hdbcli.dbapi import Error as HANAError
import threading
import concurrent.futures
import requests
import json
import logging
from openpyxl import load_workbook
from pyexcelerate import Workbook
from threading import Thread
import threading
import time
import openpyxl
import xlsxwriter
import xlrd
import pandas as pd
from concurrent.futures import ThreadPoolExecutor
import io
import math


logging.basicConfig(level=logging.DEBUG)


app = Flask(__name__)


app.secret_key = 'your_secret_key'


# Add max and min to Jinja2 environment
app.jinja_env.globals.update(max=max, min=min)



# Configure database connection (SAP HANA DB)
address = ""  # IP Address of the SAP HANA DB
port =   # Port number of the SAP HANA DB
user = ""  # Username of the SAP HANA DB
password = ""  # Password of the SAP HANA DB    


def get_db_connection():
    conn = dbapi.connect(
        address=address,
        port=port,
        user=user,
        password=password,
        compression=True
    )
    return conn

@app.route('/')
def home():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('home.html')

# Campaign1 route
@app.route('/campaign1')
def campaign1():
    if 'username' not in session:
        return redirect(url_for('login'))
    # Pass 'campaign1' to the template to show the Campaign1 content
    return render_template('home.html', campaign='campaign1')

# Campaign2 route
@app.route('/campaign2')
def campaign2():
    if 'username' not in session:
        return redirect(url_for('login'))
    # Pass 'campaign2' to the template to show the Campaign2 content
    return render_template('home.html', campaign='campaign2')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        try:
            # Call the stored procedure to validate the user
            conn = get_db_connection()
            cursor = conn.cursor()

            # Execute the stored procedure
            result = cursor.callproc('MSHAFIQ.pro_VALIDATE_USER', (email, password, 1))
            
            # Check the output parameter (login_success)
            login_success = result[2]  # The third element is the output parameter
            
            if login_success:
                # Fetch the user's role
                cursor.execute("SELECT ROLE FROM USER_CREDENTIALS WHERE EMAIL = ?", (email,))
                role = cursor.fetchone()[0]
                
                session['username'] = email
                session['role'] = role  # Store the role in the session
                
                return redirect(url_for('home'))
            else:
                flash('Invalid credentials. Please try again.')

        except dbapi.Error as e:
            # Handle the specific error related to no data found
            flash('Invalid credentials. Please try again.')
            logging.error(f"Database error: {e}")

        finally:
            cursor.close()
            conn.close()
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('username', None)
    return redirect(url_for('login'))

@app.route('/change_password', methods=['GET', 'POST'])
def change_password():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        email = session['username']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']
        
        if new_password != confirm_password:
            flash('Passwords do not match. Please try again.')
            return redirect(url_for('change_password'))
        
        # Call the stored procedure to update the password
        conn = get_db_connection()
        cursor = conn.cursor()
        
        result_message = ''
        result = cursor.callproc('MSHAFIQ.pro_UPDATE_PASSWORD', (email, new_password, result_message))
        
        cursor.close()
        conn.close()
        
        if 'successfully' in result[2]:  # Check if the password update was successful
            flash('Password updated successfully.')
            return redirect(url_for('home'))
        else:
            flash(result[2])  # Display the result message from the procedure
    
    return render_template('change_password.html')

import logging

logging.basicConfig(level=logging.DEBUG)

@app.route('/add_user', methods=['GET', 'POST'])
def add_user():
    if 'username' not in session:
        logging.debug('No username found in session. Redirecting to login.')
        return redirect(url_for('login'))

    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT ROLE FROM USER_CREDENTIALS WHERE EMAIL = ?", (session['username'],))
    user_role = cursor.fetchone()[0]
    
    logging.debug(f'User role fetched from database for {session["username"]}: {user_role}')
    
    # Convert both to uppercase to avoid case sensitivity issues
    if user_role.upper() != 'ADMIN':
        flash("You are not authorized to access this page.")
        logging.debug(f'User {session["username"]} is not admin (found role: {user_role}). Redirecting to home.')
        return redirect(url_for('home'))

    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        role = request.form['role']

        result = cursor.callproc('MSHAFIQ.pro_INSERT_USER', (email, password, role))

        logging.debug(f'User {email} added successfully.')
        
        cursor.close()
        conn.close()

        flash(f'User {email} added successfully.')
        return redirect(url_for('home'))

    cursor.close()
    conn.close()

    return render_template('add_user.html')
# Products Table
@app.route('/Products', methods=['GET', 'POST'])
def indexProducts():
    if request.method == 'POST':
        excel_file = request.files['file']
        if excel_file and excel_file.filename.endswith('.xlsx'):
            try:
                df = pd.read_excel(excel_file)
                df.columns = df.columns.str.strip().str.upper()
                
                required_columns = ['EFFECTIVESTARTDATE', 'EFFECTIVEENDDATE', 'NAME', 'PRICE', 'UNITTYPEFORPRICE']

                if not set(required_columns).issubset(df.columns):
                    flash('Error: The Excel file must have the correct column names and value types.')
                    return redirect(url_for('indexProducts'))

                df['EFFECTIVESTARTDATE'] = pd.to_datetime(df['EFFECTIVESTARTDATE'])
                df['EFFECTIVEENDDATE'] = pd.to_datetime(df['EFFECTIVEENDDATE'])

                conn = get_db_connection()
                cursor = conn.cursor()

                # Get the maximum existing PRODUCTID and start from there
                cursor.execute("SELECT MAX(PRODUCTID) FROM ZPRODUCT")
                max_product_id = cursor.fetchone()[0]
                if max_product_id is None:
                    max_product_id = 0

                error_rows = []
                current_product_id = max_product_id + 1

                for idx, row in df.iterrows():
                    try:
                        # Check if PRICE is numeric
                        if not isinstance(row['PRICE'], (int, float)):
                            raise ValueError(f"Invalid price value at row {idx + 1}")
                        
                        # Convert NAME to a string to ensure that both numeric and string values are handled
                        name = str(row['NAME'])
                        
                        # Insert/update logic here
                        cursor.execute("SELECT COUNT(*) FROM ZPRODUCT WHERE PRODUCTID = ?", (current_product_id,))
                        exists = cursor.fetchone()[0]

                        if exists:
                            cursor.execute("""
                                UPDATE ZPRODUCT
                                SET EFFECTIVESTARTDATE = ?, EFFECTIVEENDDATE = ?, CREATEDATE = ?, NAME = ?, PRICE = ?, UNITTYPEFORPRICE = ?
                                WHERE PRODUCTID = ?
                            """, (
                                row['EFFECTIVESTARTDATE'], 
                                row['EFFECTIVEENDDATE'], 
                                datetime.now(),
                                name,  # Use the converted string name
                                row['PRICE'], 
                                row['UNITTYPEFORPRICE'],
                                current_product_id
                            ))
                        else:
                            cursor.execute("""
                                INSERT INTO ZPRODUCT (PRODUCTID, EFFECTIVESTARTDATE, EFFECTIVEENDDATE, CREATEDATE, NAME, PRICE, UNITTYPEFORPRICE)
                                VALUES (?, ?, ?, ?, ?, ?, ?)
                            """, (
                                current_product_id, 
                                row['EFFECTIVESTARTDATE'], 
                                row['EFFECTIVEENDDATE'], 
                                datetime.now(),
                                name,  # Use the converted string name
                                row['PRICE'], 
                                row['UNITTYPEFORPRICE']
                            ))

                        current_product_id += 1

                    except ValueError as ve:
                        # Collect error message for invalid rows
                        error_rows.append(str(ve))
                        continue


                conn.commit()
                cursor.close()
                conn.close()

                if error_rows:
                    flash(f"You have entered a string value instead of integer/decimal in row(s): {', '.join(error_rows)}")
                else:
                    flash('Excel file data successfully processed and added/updated in the database.')

            except Exception as e:
                flash(f'Error processing the file: {str(e)}')
            return redirect(url_for('indexProducts'))
        else:
            flash('Please upload a valid .xlsx file.')
            return redirect(url_for('indexProducts'))

    search_query = request.args.get('search')
    search_field = request.args.get('field', 'PRODUCTID')
    page = request.args.get('page', 1, type=int)
    per_page = 8

    conn = get_db_connection()
    cursor = conn.cursor()

    if search_query:
        if search_field in ['PRICE']:
            try:
                search_query = float(search_query)
            except ValueError:
                flash(f"Invalid {search_field.lower()} value.")
                return redirect(url_for('indexProducts'))
            query = f"SELECT COUNT(*) FROM ZPRODUCT WHERE {search_field} = ?"
            cursor.execute(query, (search_query,))
        else:
            query = f"SELECT COUNT(*) FROM ZPRODUCT WHERE {search_field} LIKE ?"
            cursor.execute(query, (f"%{search_query}%",))

        total = cursor.fetchone()[0]

        query = f"""
        SELECT PRODUCTID, EFFECTIVESTARTDATE, EFFECTIVEENDDATE, NAME, PRICE, UNITTYPEFORPRICE
        FROM ZPRODUCT
        WHERE {search_field} LIKE ?
        ORDER BY PRODUCTID DESC
        LIMIT ? OFFSET ?
        """
        cursor.execute(query, (f"%{search_query}%", per_page, (page - 1) * per_page))
    else:
        query = "SELECT COUNT(*) FROM ZPRODUCT"
        cursor.execute(query)
        total = cursor.fetchone()[0]

        query = """
        SELECT PRODUCTID, EFFECTIVESTARTDATE, EFFECTIVEENDDATE, NAME, PRICE, UNITTYPEFORPRICE
        FROM ZPRODUCT
        ORDER BY PRODUCTID DESC
        LIMIT ? OFFSET ?
        """
        cursor.execute(query, (per_page, (page - 1) * per_page))

    data = cursor.fetchall()
    cursor.close()
    conn.close()

    total_pages = math.ceil(total / per_page)

    if 'download' in request.args:
        return export_products_to_excel(search_query, search_field)

    return render_template('indexProducts.html', data=data, page=page, total_pages=total_pages, search_query=search_query, search_field=search_field)

def export_products_to_excel(search_query=None, search_field=None):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Define the batch size for fetching rows
    batch_size = 10000
    offset = 0
    all_rows = []

    while True:
        if search_query:
            query = f"""
            SELECT PRODUCTID, EFFECTIVESTARTDATE, EFFECTIVEENDDATE, NAME, PRICE, UNITTYPEFORPRICE
            FROM ZPRODUCT
            WHERE {search_field} LIKE ?
            ORDER BY PRODUCTID DESC
            LIMIT ? OFFSET ?
            """
            params = (f"%{search_query}%", batch_size, offset)
        else:
            query = f"""
            SELECT PRODUCTID, EFFECTIVESTARTDATE, EFFECTIVEENDDATE, NAME, PRICE, UNITTYPEFORPRICE
            FROM ZPRODUCT
            ORDER BY PRODUCTID DESC
            LIMIT ? OFFSET ?
            """
            params = (batch_size, offset)

        cursor.execute(query, params)
        rows = cursor.fetchall()

        if not rows:
            break

        all_rows.extend(rows)
        offset += batch_size

    conn.close()

    # Create a DataFrame from the fetched rows
    df = pd.DataFrame(all_rows, columns=['PRODUCTID', 'EFFECTIVESTARTDATE', 'EFFECTIVEENDDATE', 'NAME', 'PRICE', 'UNITTYPEFORPRICE'])

    # Create an in-memory Excel file using xlsxwriter
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Products')

    output.seek(0)

    # Send the Excel file as a response
    return send_file(output, as_attachment=True, download_name='products.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/viewProducts/<int:productid>', methods=['GET'])
def viewProducts(productid):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Fetch the product's data
    cursor.execute("""
        SELECT PRODUCTID, EFFECTIVESTARTDATE, EFFECTIVEENDDATE, NAME, PRICE, UNITTYPEFORPRICE
        FROM ZPRODUCT 
        WHERE productid = ?
    """, (productid,))
    product_data = cursor.fetchone()
    cursor.close()
    conn.close()

    if product_data is None:
        flash('Retailer not found.', 'error')
        return redirect(url_for('indexProducts'))

    return render_template('viewProducts.html', product=product_data)

@app.route('/addProductEntry', methods=['GET', 'POST'])
def addProductEntry():
    if request.method == 'POST':
        effectivestartdate = request.form['effectivestartdate']
        effectiveenddate = request.form['effectiveenddate']
        name = request.form['name']
        price = request.form['price']
        unittypeforprice = request.form['unittypeforprice']
        
        # Validate both Name and Price as numbers (integers or decimals)
        try:
            name = float(name)  # Convert name to float
        except ValueError:
            flash('The Name field must contain a valid number (integer or decimal).')
            return redirect(url_for('addProductEntry'))

        try:
            price = float(price)  # Convert price to float
        except ValueError:
            flash('The Price field must contain a valid number (integer or decimal).')
            return redirect(url_for('addProductEntry'))

        createdate = datetime.now()

        conn = None
        cursor = None
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO ZPRODUCT (EFFECTIVESTARTDATE, EFFECTIVEENDDATE, CREATEDATE, NAME, PRICE, UNITTYPEFORPRICE)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (effectivestartdate, effectiveenddate, createdate, name, price, unittypeforprice))
            conn.commit()
            flash('New product added successfully.')
        except dbapi.Error as e:
            conn.rollback()
            flash(f'Error occurred: {str(e)}')
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()

        return redirect(url_for('indexProducts'))

    return render_template('addProductEntry.html', current_date=datetime.now().strftime('%Y-%m-%d'))


@app.route('/editProducts/<int:productid>', methods=['GET', 'POST'])
def editProducts(productid):
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        if request.method == 'POST':
            new_effectivestartdate = request.form['effectivestartdate']
            new_effectiveenddate = request.form['effectiveenddate']
            new_name = request.form['name']
            new_price = float(request.form['price'])
            new_unittypeforprice = request.form['unittypeforprice']
           
            new_createdate = datetime.now()

            cursor.execute("""
            UPDATE ZPRODUCT
            SET EFFECTIVESTARTDATE = ?, EFFECTIVEENDDATE = ?, CREATEDATE = ?, NAME = ?, PRICE = ?, UNITTYPEFORPRICE = ?
            WHERE PRODUCTID = ?
            """, (new_effectivestartdate, new_effectiveenddate, new_createdate, new_name, new_price, new_unittypeforprice, productid))
            
            conn.commit()
            flash('Entry successfully updated.')
            return redirect(url_for('indexProducts'))
        
        cursor.execute("SELECT PRODUCTID, EFFECTIVESTARTDATE, EFFECTIVEENDDATE, NAME, PRICE, UNITTYPEFORPRICE FROM ZPRODUCT WHERE PRODUCTID = ?", (productid,))
        entry = cursor.fetchone()

    except dbapi.Error as e:
        flash(f'Error occurred: {str(e)}')
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    return render_template('editProducts.html', entry=entry)


@app.route('/deleteProducts/<productid>', methods=['POST'])
def deleteProducts(productid):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM ZPRODUCT WHERE PRODUCTID = ?", (productid,))
    conn.commit()
    cursor.close()
    conn.close()
    flash('Entry successfully deleted.')
    return redirect(url_for('indexProducts'))



@app.route('/displayProducts')
def displayProducts():
        # Handle search and pagination for Products
    search_query = request.args.get('search')  # Extract the search query from form input
    search_field = request.args.get('field', 'PRODUCTID')  # Extract the search field, default to 'PRODUCTID'
    page = request.args.get('page', 1, type=int)
    per_page = 8

    conn = get_db_connection()
    cursor = conn.cursor()

    if search_query:
        # Build the SQL query based on the selected field
        if search_field in ['PRICE']:
            # Convert search_query to a numeric type for price or cost searches
            try:
                search_query = float(search_query)
            except ValueError:
                flash(f"Invalid {search_field.lower()} value.")
                return redirect(url_for('indexProducts'))
            query = f"SELECT COUNT(*) FROM ZPRODUCT WHERE {search_field} = ?"
            cursor.execute(query, (search_query,))
        else:
            # Handle text-based search (e.g., NAME)
            query = f"SELECT COUNT(*) FROM ZPRODUCT WHERE {search_field} LIKE ?"
            cursor.execute(query, (f"%{search_query}%",))

        total = cursor.fetchone()[0]

        query = f"""
        SELECT PRODUCTID, EFFECTIVESTARTDATE, EFFECTIVEENDDATE, NAME, PRICE, UNITTYPEFORPRICE
        FROM ZPRODUCT
        WHERE {search_field} LIKE ?
        ORDER BY PRODUCTID ASC
        LIMIT ? OFFSET ?
        """
        cursor.execute(query, (f"%{search_query}%", per_page, (page - 1) * per_page))
    else:
        # No search query, return all products
        query = "SELECT COUNT(*) FROM ZPRODUCT"
        cursor.execute(query)
        total = cursor.fetchone()[0]

        query = """
        SELECT PRODUCTID, EFFECTIVESTARTDATE, EFFECTIVEENDDATE, NAME, PRICE, UNITTYPEFORPRICE
        FROM ZPRODUCT
        ORDER BY PRODUCTID ASC
        LIMIT ? OFFSET ?
        """
        cursor.execute(query, (per_page, (page - 1) * per_page))

    data = cursor.fetchall()
    cursor.close()
    conn.close()

    total_pages = math.ceil(total / per_page)
  

    return render_template('displayProducts.html', data=data, page=page, total_pages=total_pages, search_query=search_query)

#For Retailer Table

def stream_excel(data_iterator):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet('Retailers')

    column_mapping = {
        'RETAILERID': 'RetailerID',
        'EFFECTIVESTARTDATE': 'StartDate',
        'EFFECTIVEENDDATE': 'EndDate',
        'GENERICATTRIBUTE1': 'NTN',
        'GENERICATTRIBUTE2': 'RetailerMSISDN',
        'GENERICATTRIBUTE3': 'Zone',
        'GENERICATTRIBUTE4': 'CNIC',
        'GENERICATTRIBUTE5': 'FranchiseID',
        'GENERICATTRIBUTE6': 'City',
        'GENERICATTRIBUTE7': 'Region',
        'GENERICATTRIBUTE8': 'ContactPerson',
        'GENERICATTRIBUTE9': 'PartnerPostalAddress',
        'GENERICATTRIBUTE10': 'RSOCode',
        'GENERICATTRIBUTE11': 'TaxCombination',
        'GENERICATTRIBUTE12': 'Email Address',
        'GENERICATTRIBUTE13': 'CategoryName',
        'GENERICATTRIBUTE14': 'Lattitude',
        'GENERICATTRIBUTE15': 'Longitude',
        'GENERICATTRIBUTE16': 'ContactNumber',
        'GENERICNUMBER1': 'WHTTaxRate',
        'GENERICNUMBER2': 'GSTTaxRate',
        'GENERICNUMBER3': 'GSTWithheld',
        'GENERICNUMBER4': 'AggTaxRate',
        'GENERICNUMBER5': 'PartnerTypeID',
        'GENERICNUMBER6': 'PartnerClassID',
        'GENERICDATE1': 'CreationDate',
        'GENERICDATE2': 'UpdateDate',
        'GENERICBOOLEAN1': 'TaxExemptedFlag',
        'GENERICBOOLEAN2': 'Status',
        'NAME': 'Name',
        'DESCRIPTION': 'Description',
        'CREATEDATE': 'CreateDate'
    }

    # Write the header
    for col_num, header in enumerate(column_mapping.values()):
        worksheet.write(0, col_num, header)

    # Write data rows
    row_num = 1
    for data_chunk in data_iterator:
        for row in data_chunk:
            for col_num, (db_field, _) in enumerate(column_mapping.items()):
                worksheet.write(row_num, col_num, row[db_field])
            row_num += 1

    workbook.close()
    output.seek(0)
    yield output.read()

@app.route('/Retailers', methods=['GET'])
def indexRetailers():
    search_query = request.args.get('search')
    search_field = request.args.get('field', 'RETAILERID').upper()
    download = request.args.get('download', False)
    chunk_size = 10000  # Fetch 10,000 rows at a time
    page = request.args.get('page', 1, type=int)
    per_page = 8

    conn = get_db_connection()
    cursor = conn.cursor()

    if download:
        # Modify the query to fetch all data without LIMIT/OFFSET when downloading
        if search_query:
            query = f"""
            SELECT RETAILERID, EFFECTIVESTARTDATE, EFFECTIVEENDDATE,
                            GENERICATTRIBUTE1, GENERICATTRIBUTE2, GENERICATTRIBUTE3, GENERICATTRIBUTE4, 
                            GENERICATTRIBUTE5, GENERICATTRIBUTE6, GENERICATTRIBUTE7, GENERICATTRIBUTE8, 
                            GENERICATTRIBUTE9, GENERICATTRIBUTE10, GENERICATTRIBUTE11, GENERICATTRIBUTE12, 
                            GENERICATTRIBUTE13, GENERICATTRIBUTE14, GENERICATTRIBUTE15, GENERICATTRIBUTE16, 
                            GENERICNUMBER1, GENERICNUMBER2, GENERICNUMBER3, GENERICNUMBER4, GENERICNUMBER5, 
                            GENERICNUMBER6, GENERICDATE1, GENERICDATE2, GENERICBOOLEAN1, GENERICBOOLEAN2, 
                            NAME, DESCRIPTION, CREATEDATE
            FROM ZRETAILER
            WHERE {search_field} LIKE ?
            ORDER BY RETAILERID DESC
            """
            cursor.execute(query, (f"%{search_query}%",))
        else:
            query = """
            SELECT RETAILERID, EFFECTIVESTARTDATE, EFFECTIVEENDDATE,
                            GENERICATTRIBUTE1, GENERICATTRIBUTE2, GENERICATTRIBUTE3, GENERICATTRIBUTE4, 
                            GENERICATTRIBUTE5, GENERICATTRIBUTE6, GENERICATTRIBUTE7, GENERICATTRIBUTE8, 
                            GENERICATTRIBUTE9, GENERICATTRIBUTE10, GENERICATTRIBUTE11, GENERICATTRIBUTE12, 
                            GENERICATTRIBUTE13, GENERICATTRIBUTE14, GENERICATTRIBUTE15, GENERICATTRIBUTE16, 
                            GENERICNUMBER1, GENERICNUMBER2, GENERICNUMBER3, GENERICNUMBER4, GENERICNUMBER5, 
                            GENERICNUMBER6, GENERICDATE1, GENERICDATE2, GENERICBOOLEAN1, GENERICBOOLEAN2, 
                            NAME, DESCRIPTION, CREATEDATE
            FROM ZRETAILER
            ORDER BY RETAILERID DESC
            """
            cursor.execute(query)

        # Fetch data in chunks and yield it to the Excel generator
        def data_iterator():
            while True:
                data_chunk = cursor.fetchmany(chunk_size)
                if not data_chunk:
                    break
                yield data_chunk

        # Stream the Excel file
        headers = {
            'Content-Disposition': 'attachment; filename="retailers_data.xlsx"',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        return Response(stream_with_context(stream_excel(data_iterator())), headers=headers)

    else:
        # Paginated query for non-download requests
        if search_query:
            query = f"""
            SELECT RETAILERID, EFFECTIVESTARTDATE, EFFECTIVEENDDATE,
                            GENERICATTRIBUTE1, GENERICATTRIBUTE2, GENERICATTRIBUTE3, GENERICATTRIBUTE4, 
                            GENERICATTRIBUTE5, GENERICATTRIBUTE6, GENERICATTRIBUTE7, GENERICATTRIBUTE8, 
                            GENERICATTRIBUTE9, GENERICATTRIBUTE10, GENERICATTRIBUTE11, GENERICATTRIBUTE12, 
                            GENERICATTRIBUTE13, GENERICATTRIBUTE14, GENERICATTRIBUTE15, GENERICATTRIBUTE16, 
                            GENERICNUMBER1, GENERICNUMBER2, GENERICNUMBER3, GENERICNUMBER4, GENERICNUMBER5, 
                            GENERICNUMBER6, GENERICDATE1, GENERICDATE2, GENERICBOOLEAN1, GENERICBOOLEAN2, 
                            NAME, DESCRIPTION, CREATEDATE
            FROM ZRETAILER
            WHERE {search_field} LIKE ?
            ORDER BY RETAILERID DESC
            LIMIT ? OFFSET ?
            """
            cursor.execute(query, (f"%{search_query}%", per_page, (page - 1) * per_page))
        else:
            query = """
            SELECT RETAILERID, EFFECTIVESTARTDATE, EFFECTIVEENDDATE,
                            GENERICATTRIBUTE1, GENERICATTRIBUTE2, GENERICATTRIBUTE3, GENERICATTRIBUTE4, 
                            GENERICATTRIBUTE5, GENERICATTRIBUTE6, GENERICATTRIBUTE7, GENERICATTRIBUTE8, 
                            GENERICATTRIBUTE9, GENERICATTRIBUTE10, GENERICATTRIBUTE11, GENERICATTRIBUTE12, 
                            GENERICATTRIBUTE13, GENERICATTRIBUTE14, GENERICATTRIBUTE15, GENERICATTRIBUTE16, 
                            GENERICNUMBER1, GENERICNUMBER2, GENERICNUMBER3, GENERICNUMBER4, GENERICNUMBER5, 
                            GENERICNUMBER6, GENERICDATE1, GENERICDATE2, GENERICBOOLEAN1, GENERICBOOLEAN2, 
                            NAME, DESCRIPTION, CREATEDATE
            FROM ZRETAILER
            ORDER BY RETAILERID DESC
            LIMIT ? OFFSET ?
            """
            cursor.execute(query, (per_page, (page - 1) * per_page))

        data = cursor.fetchall()

        # Get the total number of records for pagination
        cursor.execute('SELECT COUNT(*) FROM ZRETAILER')
        total = cursor.fetchone()[0]
        total_pages = (total + per_page - 1) // per_page

        cursor.close()
        conn.close()

        return render_template('indexRetailers.html', data=data, page=page, total_pages=total_pages, search_query=search_query, search_field=search_field)


@app.route('/addRetailerEntry', methods=['GET', 'POST'])
def addRetailerEntry():
    if request.method == 'POST':
        retailerid = request.form['retailerid']
        
        # Check if retailerid already exists in the database
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT 1 FROM ZRETAILER WHERE retailerid = ?", (retailerid,))
        existing_retailer = cursor.fetchone()

        if existing_retailer:
            # Retailer ID already exists, flash a message and redirect
            flash(f'Retailer ID {retailerid} already exists.', 'error')
            cursor.close()
            conn.close()
            return redirect(url_for('indexRetailers'))
        
        # Retrieve other form fields
        genericattribute1 = request.form['genericattribute1']
        genericattribute2 = request.form['genericattribute2']
        genericattribute3 = request.form['genericattribute3']
        genericattribute4 = request.form['genericattribute4']
        genericattribute5 = request.form['genericattribute5']
        genericattribute6 = request.form['genericattribute6']
        genericattribute7 = request.form['genericattribute7']
        genericattribute8 = request.form['genericattribute8']
        genericattribute9 = request.form['genericattribute9']
        genericattribute10 = request.form['genericattribute10']
        genericattribute11 = request.form['genericattribute11']
        genericattribute12 = request.form['genericattribute12']
        genericattribute13 = request.form['genericattribute13']
        genericattribute14 = request.form['genericattribute14']
        genericattribute15 = request.form['genericattribute15']
        genericattribute16 = request.form['genericattribute16']
        genericnumber1 = request.form['genericnumber1']
        genericnumber2 = request.form['genericnumber2']
        genericnumber3 = request.form['genericnumber3']
        genericnumber4 = request.form['genericnumber4']
        genericnumber5 = request.form['genericnumber5']
        genericnumber6 = request.form['genericnumber6']
        genericboolean1 = request.form['genericboolean1']
        genericboolean2 = request.form['genericboolean2']
        name = request.form['name']
        description = request.form['description']
        effectivestartdate = request.form['effectivestartdate']
        effectiveenddate = request.form['effectiveenddate']
        genericdate1 = request.form['genericdate1']
        genericdate2 = datetime.now()

        # Automatically setting dates
        createdate = datetime.now()
        islast = 1

        # Insert new retailer entry
        cursor.execute("""
            INSERT INTO ZRETAILER (retailerid, effectivestartdate, effectiveenddate, genericattribute1, genericattribute2, genericattribute3, genericattribute4, genericattribute5, genericattribute6, genericattribute7, genericattribute8, genericattribute9, genericattribute10, genericattribute11, genericattribute12, genericattribute13, genericattribute14, genericattribute15, genericattribute16, genericnumber1, genericnumber2, genericnumber3, genericnumber4, genericnumber5, genericnumber6, genericdate1, genericdate2, genericboolean1, genericboolean2, name, description, islast, createdate)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (retailerid, effectivestartdate, effectiveenddate, genericattribute1, genericattribute2, genericattribute3, genericattribute4, genericattribute5, genericattribute6, genericattribute7, genericattribute8, genericattribute9, genericattribute10, genericattribute11, genericattribute12, genericattribute13, genericattribute14, genericattribute15, genericattribute16, genericnumber1, genericnumber2, genericnumber3, genericnumber4, genericnumber5, genericnumber6, genericdate1, genericdate2, genericboolean1, genericboolean2, name, description, islast, createdate))
        conn.commit()

        # Close the cursor and connection
        cursor.close()
        conn.close()

        flash(f'New retailer {retailerid} entry added successfully.')
        return redirect(url_for('indexRetailers'))

    return render_template('addRetailerEntry.html')

@app.route('/editRetailers/<retailerid>', methods=['GET', 'POST'])
def editRetailers(retailerid):
    conn = get_db_connection()
    cursor = conn.cursor()
    if request.method == 'POST':
        new_genericattribute1 = request.form['genericattribute1']
        new_genericattribute2 = request.form['genericattribute2']
        new_genericattribute3 = request.form['genericattribute3']
        new_genericattribute4 = request.form['genericattribute4']
        new_genericattribute5 = request.form['genericattribute5']
        new_genericattribute6 = request.form['genericattribute6']
        new_genericattribute7 = request.form['genericattribute7']
        new_genericattribute8 = request.form['genericattribute8']
        new_genericattribute9 = request.form['genericattribute9']
        new_genericattribute10 = request.form['genericattribute10']
        new_genericattribute11 = request.form['genericattribute11']
        new_genericattribute12 = request.form['genericattribute12']
        new_genericattribute13 = request.form['genericattribute13']
        new_genericattribute14 = request.form['genericattribute14']
        new_genericattribute15 = request.form['genericattribute15']
        new_genericattribute16 = request.form['genericattribute16']
        new_genericnumber1 = request.form['genericnumber1']
        new_genericnumber2 = request.form['genericnumber2']
        new_genericnumber3 = request.form['genericnumber3']
        new_genericnumber4 = request.form['genericnumber4']
        new_genericnumber5 = request.form['genericnumber5']
        new_genericnumber6 = request.form['genericnumber6']
        new_genericboolean1 = request.form['genericboolean1']
        new_genericboolean2 = request.form['genericboolean2']
        new_name = request.form['name']
        new_description = request.form['description']
    
        new_effectivestartdate = request.form['effectivestartdate']
        new_effectiveenddate = request.form['effectiveenddate']
        new_genericdate1 = request.form['genericdate1']
        new_genericdate2 = datetime.now()

        # Automatically updating dates
      
        new_createdate = datetime.now()
        


        new_islast = 1

        cursor.execute("""
        UPDATE ZRETAILER 
        SET effectivestartdate = ?, effectiveenddate = ?, genericattribute1 = ?, genericattribute2 = ?, genericattribute3 = ?, genericattribute4 = ?, genericattribute5 = ?, genericattribute6 = ?, genericattribute7 = ?, genericattribute8 = ?, genericattribute9 = ?, genericattribute10 = ?, genericattribute11 = ?, genericattribute12 = ?, genericattribute13 = ?, genericattribute14 = ?, genericattribute15 = ?, genericattribute16 = ?, genericnumber1 = ?, genericnumber2 = ?, genericnumber3 = ?, genericnumber4 = ?, genericnumber5 = ?, genericnumber6 = ?, genericdate1 = ?, genericdate2 = ?, genericboolean1 = ?, genericboolean2 = ?, name = ?, description = ?, islast = ?, createdate = ?
        WHERE retailerid = ?
        """, (new_effectivestartdate, new_effectiveenddate, new_genericattribute1, new_genericattribute2, new_genericattribute3, new_genericattribute4, new_genericattribute5, new_genericattribute6, new_genericattribute7, new_genericattribute8, new_genericattribute9, new_genericattribute10, new_genericattribute11, new_genericattribute12, new_genericattribute13, new_genericattribute14, new_genericattribute15, new_genericattribute16, new_genericnumber1, new_genericnumber2, new_genericnumber3, new_genericnumber4, new_genericnumber5, new_genericnumber6, new_genericdate1, new_genericdate2, new_genericboolean1, new_genericboolean2, new_name, new_description, new_islast, new_createdate, retailerid))
        
        conn.commit()
        cursor.close()
        conn.close()
        flash('Entry successfully updated.')
        return redirect(url_for('indexRetailers'))
    
    cursor.execute("""
    SELECT retailerid, effectivestartdate, effectiveenddate, genericattribute1, genericattribute2, genericattribute3, genericattribute4, genericattribute5, genericattribute6, genericattribute7, genericattribute8, genericattribute9, genericattribute10, genericattribute11, genericattribute12, genericattribute13, genericattribute14, genericattribute15, genericattribute16, genericnumber1, genericnumber2, genericnumber3, genericnumber4, genericnumber5, genericnumber6, genericdate1, genericdate2, genericboolean1, genericboolean2, name, description, createdate FROM ZRETAILER WHERE retailerid = ?""", (retailerid,))
    entry = cursor.fetchone()
    cursor.close()
    conn.close()
    return render_template('editRetailers.html', entry=entry)

@app.route('/deleteRetailers/<retailerid>', methods=['POST'])
def deleteRetailers(retailerid):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM ZRETAILER WHERE retailerid = ?", (retailerid,))
    conn.commit()
    cursor.close()
    conn.close()
    flash('Entry successfully deleted.')
    return redirect(url_for('indexRetailers'))

@app.route('/viewRetailer/<retailerid>', methods=['GET'])
def viewRetailer(retailerid):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Fetch the retailer's data
    cursor.execute("""
        SELECT retailerid, effectivestartdate, effectiveenddate, genericattribute1, genericattribute2, genericattribute3, genericattribute4, genericattribute5, genericattribute6, genericattribute7, genericattribute8, genericattribute9, genericattribute10, genericattribute11, genericattribute12, genericattribute13, genericattribute14, genericattribute15, genericattribute16, genericnumber1, genericnumber2, genericnumber3, genericnumber4, genericnumber5, genericnumber6, genericdate1, genericdate2, genericboolean1, genericboolean2, name, description, createdate 
        FROM ZRETAILER 
        WHERE retailerid = ?
    """, (retailerid,))
    retailer_data = cursor.fetchone()
    cursor.close()
    conn.close()

    if retailer_data is None:
        flash('Retailer not found.', 'error')
        return redirect(url_for('indexRetailers'))

    return render_template('viewRetailer.html', retailer=retailer_data)



# Retailer Error Table: ZRETAILER_ERROR
@app.route('/RetailersError')
def indexRetailersError():
    search_query = request.args.get('search')
    page = request.args.get('page', 1, type=int)
    per_page = 8

    conn = get_db_connection()
    cursor = conn.cursor()

    if search_query:
        query = "SELECT COUNT(*) FROM ZRETAILER WHERE RETAILERID = ?"
        cursor.execute(query, (search_query,))
        total = cursor.fetchone()[0]

        query = """
        SELECT retailerid, effectivestartdate, effectiveenddate, genericattribute1, genericattribute2, genericattribute3, genericattribute4, genericattribute5, genericattribute6, genericattribute7, genericattribute8, genericattribute9, genericattribute10, genericattribute11, genericattribute12, genericattribute13, genericattribute14, genericattribute15, genericattribute16, genericnumber1, genericnumber2, genericnumber3, genericnumber4, genericnumber5, genericnumber6, genericdate1, genericdate2, genericdate3, genericboolean1, genericboolean2, name, description, islast, createdate
        FROM ZRETAILER
        WHERE retailerid = ?
        ORDER BY retailerid ASC
        LIMIT ? OFFSET ?
        """
        cursor.execute(query, (search_query, per_page, (page - 1) * per_page))
    else:
        query = "SELECT COUNT(*) FROM ZRETAILER"
        cursor.execute(query)
        total = cursor.fetchone()[0]

        query = """
        SELECT retailerid, effectivestartdate, effectiveenddate, genericattribute1, genericattribute2, genericattribute3, genericattribute4, genericattribute5, genericattribute6, genericattribute7, genericattribute8, genericattribute9, genericattribute10, genericattribute11, genericattribute12, genericattribute13, genericattribute14, genericattribute15, genericattribute16, genericnumber1, genericnumber2, genericnumber3, genericnumber4, genericnumber5, genericnumber6, genericdate1, genericdate2, genericdate3, genericboolean1, genericboolean2, name, description, islast, createdate
        FROM ZRETAILER
        ORDER BY Retailerid ASC
        LIMIT ? OFFSET ?
        """
        cursor.execute(query, (per_page, (page - 1) * per_page))

    data = cursor.fetchall()
    cursor.close()
    conn.close()

    total_pages = math.ceil(total / per_page)
    return render_template('indexRetailersError.html', data=data, page=page, total_pages=total_pages, search_query=search_query)


@app.route('/editRetailersError/<retailerid>', methods=['GET', 'POST'])
def editRetailersError(retailerid):
    conn = get_db_connection()
    cursor = conn.cursor()
    if request.method == 'POST':
        new_effectivestartdate = request.form['effectivestartdate']
        new_effectiveenddate = request.form['effectiveenddate']
        new_genericattribute1 = request.form['genericattribute1']
        new_genericattribute2 = request.form['genericattribute2']
        new_genericattribute3 = request.form['genericattribute3']
        new_genericattribute4 = request.form['genericattribute4']
        new_genericattribute5 = request.form['genericattribute5']
        new_genericattribute6 = request.form['genericattribute6']
        new_genericattribute7 = request.form['genericattribute7']
        new_genericattribute8 = request.form['genericattribute8']
        new_genericattribute9 = request.form['genericattribute9']
        new_genericattribute10 = request.form['genericattribute10']
        new_genericattribute11 = request.form['genericattribute11']
        new_genericattribute12 = request.form['genericattribute12']
        new_genericattribute13 = request.form['genericattribute13']
        new_genericattribute14 = request.form['genericattribute14']
        new_genericattribute15 = request.form['genericattribute15']
        new_genericattribute16 = request.form['genericattribute16']
        new_genericnumber1 = request.form['genericnumber1']
        new_genericnumber2 = request.form['genericnumber2']
        new_genericnumber3 = request.form['genericnumber3']
        new_genericnumber4 = request.form['genericnumber4']
        new_genericnumber5 = request.form['genericnumber5']
        new_genericnumber6 = request.form['genericnumber6']
        new_genericdate1 = request.form['genericdate1']
        new_genericdate2 = request.form['genericdate2']
        new_genericdate3 = request.form['genericdate3']
        new_genericboolean1 = request.form['genericboolean1']
        new_genericboolean2 = request.form['genericboolean2']
        new_name = request.form['name']
        new_description = request.form['description']
        new_islast = request.form['islast']
        new_createdate = request.form['createdate']
        
        cursor.execute("""
        UPDATE ZRETAILER 
        SET effectivestartdate = ?, effectiveenddate = ?, genericattribute1 = ?, genericattribute2 = ?, genericattribute3 = ?, genericattribute4 = ?, genericattribute5 = ?, genericattribute6 = ?, genericattribute7 = ?, genericattribute8 = ?, genericattribute9 = ?, genericattribute10 = ?, genericattribute11 = ?, genericattribute12 = ?, genericattribute13 = ?, genericattribute14 = ?, genericattribute15 = ?, genericattribute16 = ?, genericnumber1 = ?, genericnumber2 = ?, genericnumber3 = ?, genericnumber4 = ?, genericnumber5 = ?, genericnumber6 = ?, genericdate1 = ?, genericdate2 = ?, genericdate3 = ?, genericboolean1 = ?, genericboolean2 = ?, name = ?, description = ?, islast = ?, createdate = ?
        WHERE retailerid = ?
        """, (new_effectivestartdate, new_effectiveenddate, new_genericattribute1, new_genericattribute2, new_genericattribute3, new_genericattribute4, new_genericattribute5, new_genericattribute6, new_genericattribute7, new_genericattribute8, new_genericattribute9, new_genericattribute10, new_genericattribute11, new_genericattribute12, new_genericattribute13, new_genericattribute14, new_genericattribute15, new_genericattribute16, new_genericnumber1, new_genericnumber2, new_genericnumber3, new_genericnumber4, new_genericnumber5, new_genericnumber6, new_genericdate1, new_genericdate2, new_genericdate3, new_genericboolean1, new_genericboolean2, new_name, new_description, new_islast, new_createdate, retailerid))
        
        conn.commit()
        cursor.close()
        conn.close()
        flash('Entry successfully updated.')
        return redirect(url_for('indexRetailersError'))
    
    cursor.execute("SELECT retailerid, effectivestartdate, effectiveenddate, genericattribute1, genericattribute2, genericattribute3, genericattribute4, genericattribute5, genericattribute6, genericattribute7, genericattribute8, genericattribute9, genericattribute10, genericattribute11, genericattribute12, genericattribute13, genericattribute14, genericattribute15, genericattribute16, genericnumber1, genericnumber2, genericnumber3, genericnumber4, genericnumber5, genericnumber6, genericdate1, genericdate2, genericdate3, genericboolean1, genericboolean2, name, description, islast, createdate FROM ZRETAILER WHERE retailerid = ?", (retailerid,))
    entry = cursor.fetchone()
    cursor.close()
    conn.close()
    return render_template('editRetailersError.html', entry=entry)

@app.route('/deleteRetailersError/<retailerid>', methods=['POST'])
def deleteRetailersError(retailerid):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM ZRETAILER WHERE retailerid = ?", (retailerid,))
    conn.commit()
    cursor.close()
    conn.close()
    flash('Entry successfully deleted.')
    return redirect(url_for('indexRetailersError'))








#Sales Transaction Table:


@app.route('/SalesTransactions')
def indexSalesTransactions():
    search_query = request.args.get('search')
    page = request.args.get('page', 1, type=int)
    per_page = 8

    conn = get_db_connection()
    cursor = conn.cursor()

    if search_query:
        query = "SELECT COUNT(*) FROM ZSALESTRANSACTION WHERE salestransactionseq = ?"
        cursor.execute(query, (search_query,))
        total = cursor.fetchone()[0]

        query = """
        SELECT salestransactionseq, channel, productid, genericattribute1, genericattribute2, genericattribute3, genericattribute12, linenumber, sublinenumber, eventtype, origintypeid, compensationdate, isrunnable, preadjustedvalue, unittypeforpreadjustedvalue, value, unittypeforvalue, modificationdate
        FROM ZSALESTRANSACTION
        WHERE salestransactionseq = ?
        ORDER BY salestransactionseq ASC
        LIMIT ? OFFSET ?
        """
        cursor.execute(query, (search_query, per_page, (page - 1) * per_page))
    else:
        query = "SELECT COUNT(*) FROM ZSALESTRANSACTION"
        cursor.execute(query)
        total = cursor.fetchone()[0]

        query = """
        SELECT salestransactionseq, channel, productid, genericattribute1, genericattribute2, genericattribute3, genericattribute12, linenumber, sublinenumber, eventtype, origintypeid, compensationdate, isrunnable, preadjustedvalue, unittypeforpreadjustedvalue, value, unittypeforvalue, modificationdate
        FROM ZSALESTRANSACTION
        ORDER BY salestransactionseq ASC
        LIMIT ? OFFSET ?
        """
        cursor.execute(query, (per_page, (page - 1) * per_page))

    data = cursor.fetchall()
    cursor.close()
    conn.close()

    total_pages = math.ceil(total / per_page)
    return render_template('indexSalesTransactions.html', data=data, page=page, total_pages=total_pages, search_query=search_query)

@app.route('/addSalesEntry', methods=['GET', 'POST'])
def addSalesEntry():
    if request.method == 'POST':
        salestransactionseq = request.form['salestransactionseq']
        orderid = request.form['orderid']
        channel = request.form['channel']
        productid = request.form['productid']
        genericattribute1 = request.form['genericattribute1']
        genericattribute2 = request.form['genericattribute2']
        genericattribute3 = request.form['genericattribute3']
        genericattribute12 = request.form['genericattribute12']
        linenumber = request.form['linenumber']
        sublinenumber = request.form['sublinenumber']
        eventtype = request.form['eventtype']  
        origintypeid = request.form['origintypeid'] 
        compensationdate = request.form['compensationdate']  
        isrunnable = request.form['isrunnable']
        preadjustedvalue = request.form['preadjustedvalue']
        unittypeforpreadjustedvalue = request.form['unittypeforpreadjustedvalue']
        value = request.form['value']
        unittypeforvalue = request.form['unittypeforvalue'] 
        modificationdate = request.form['modificationdate'] 

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO ZSALESTRANSACTION (salestransactionseq, orderid, channel, productid, genericattribute1, genericattribute2, genericattribute3, genericattribute12, linenumber, sublinenumber, eventtype, origintypeid, compensationdate, isrunnable, preadjustedvalue, unittypeforpreadjustedvalue, value, unittypeforvalue, modificationdate)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (salestransactionseq, orderid, channel, productid, genericattribute1, genericattribute2, genericattribute3, genericattribute12, linenumber, sublinenumber, eventtype, origintypeid, compensationdate, isrunnable, preadjustedvalue, unittypeforpreadjustedvalue, value, unittypeforvalue, modificationdate))
        conn.commit()
        cursor.close()
        conn.close()

        flash('New entry added successfully.')
        return redirect(url_for('indexSalesTransactions'))

    return render_template('addSalesEntry.html')




@app.route('/edit/<int:salestransactionseq>', methods=['GET', 'POST'])
def editSalesTransactions(salestransactionseq):
    conn = get_db_connection()
    cursor = conn.cursor()
    if request.method == 'POST':
        new_channel = request.form['channel']
        new_productid = request.form['productid']
        new_genericattribute1 = request.form['genericattribute1']
        new_genericattribute2 = request.form['genericattribute2']
        new_genericattribute3 = request.form['genericattribute3']
        new_genericattribute12 = request.form['genericattribute12']
        new_linenumber = request.form['linenumber']
        new_sublinenumber = request.form['sublinenumber']
        new_eventtype = request.form['eventtype']
        new_origintypeid = request.form['origintypeid']
        new_compensationdate = request.form['compensationdate']
        new_isrunnable = request.form['isrunnable']
        new_preadjustedvalue = request.form['preadjustedvalue']
        new_unittypeforpreadjustedvalue = request.form['unittypeforpreadjustedvalue']
        new_value = request.form['value']
        new_unittypeforvalue = request.form['unittypeforvalue']
        new_modificationdate = request.form['modificationdate']

        

        
        cursor.execute("""
        UPDATE ZSALESTRANSACTION 
        SET channel = ?, productid = ?, genericattribute1 = ?, genericattribute2 = ?, genericattribute3 = ?, genericattribute12 = ?, linenumber = ?, sublinenumber = ?, eventtype = ?, origintypeid = ?, compensationdate = ?, isrunnable = ?, preadjustedvalue = ?, unittypeforpreadjustedvalue = ?, value = ?, unittypeforvalue = ?, modificationdate = ?
        WHERE salestransactionseq = ?
        """, (new_channel, new_productid, new_genericattribute1, new_genericattribute2, new_genericattribute3, new_genericattribute12, new_linenumber, new_sublinenumber, new_eventtype, new_origintypeid, new_compensationdate, new_isrunnable, new_preadjustedvalue, new_unittypeforpreadjustedvalue, new_value, new_unittypeforvalue, new_modificationdate, salestransactionseq))
        
        conn.commit()
        cursor.close()
        conn.close()
        flash('Entry successfully updated.')
        return redirect(url_for('indexSalesTransactions'))
    
    cursor.execute("SELECT salestransactionseq, channel, productid, genericattribute1, genericattribute2, genericattribute3, genericattribute12, linenumber, sublinenumber, eventtype, origintypeid, compensationdate, isrunnable, preadjustedvalue, unittypeforpreadjustedvalue, value, unittypeforvalue, modificationdate FROM ZSALESTRANSACTION WHERE salestransactionseq = ?", (salestransactionseq,))
    entry = cursor.fetchone()
    cursor.close()
    conn.close()
    return render_template('editSalesTransactions.html', entry=entry)

@app.route('/delete/<int:salestransactionseq>', methods=['POST'])
def deleteSalesTransactions(salestransactionseq):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM ZSALESTRANSACTION WHERE salestransactionseq = ?", (salestransactionseq,))
    conn.commit()
    cursor.close()
    conn.close()
    flash('Entry successfully deleted.')
    return redirect(url_for('indexSalesTransactions'))



# History Sales Transaction


@app.route('/historySalesTransactions')
def indexhistorySalesTransactions():
    search_query = request.args.get('search')
    search_field = request.args.get('field', 'salestransactionseq')  # Default search field
    page = request.args.get('page', 1, type=int)
    per_page = 8

    # Get download parameters
    download = request.args.get('download', False)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    conn = get_db_connection()
    cursor = conn.cursor()

    # List of valid search fields
    valid_fields = [
        'salestransactionseq', 'comments', 'compensationdate', 'eventtype', 'genericattribute1', 'genericattribute2',
        'genericattribute3', 'genericattribute6', 'genericattribute12', 'genericnumber1', 'genericnumber2',
        'genericnumber3', 'genericnumber4', 'genericnumber5', 'linenumber', 'orderid', 'origintypeid',
        'productdescription', 'productid', 'productname', 'sublinenumber', 'value', 'genericattribute27',
        'genericattribute28', 'genericattribute29', 'genericattribute30', 'genericattribute22', 'genericattribute23',
        'genericattribute24', 'genericattribute31', 'genericboolean1', 'genericboolean2'
    ]

    # Query for counting total records based on search criteria
    if search_query and search_field in valid_fields:
        count_query = f"SELECT COUNT(*) FROM ZSALESTRANSACTION_LIST WHERE {search_field} LIKE ?"
        params = (f"%{search_query}%",)
        cursor.execute(count_query, params)
        total = cursor.fetchone()[0]

        # Data query to retrieve paginated results based on search query
        data_query = f"""
            SELECT salestransactionseq, comments, compensationdate, eventtype, genericattribute1, genericattribute2,
                   genericattribute3, genericattribute6, genericattribute12, genericnumber1, genericnumber2,
                   genericnumber3, genericnumber4, genericnumber5, linenumber, orderid, origintypeid,
                   productdescription, productid, productname, sublinenumber, value, genericattribute27,
                   genericattribute28, genericattribute29, genericattribute30, genericattribute22, genericattribute23,
                   genericattribute24, genericattribute31, genericboolean1, genericboolean2
            FROM ZSALESTRANSACTION_LIST
            WHERE {search_field} LIKE ?
            ORDER BY salestransactionseq DESC
            LIMIT ? OFFSET ?
        """
        cursor.execute(data_query, (f"%{search_query}%", per_page, (page - 1) * per_page))
    else:
        # Query for counting total records without search query
        count_query = "SELECT COUNT(*) FROM ZSALESTRANSACTION_LIST"
        cursor.execute(count_query)
        total = cursor.fetchone()[0]

        # Data query to retrieve paginated results
        data_query = """
            SELECT salestransactionseq, comments, compensationdate, eventtype, genericattribute1, genericattribute2,
                   genericattribute3, genericattribute6, genericattribute12, genericnumber1, genericnumber2,
                   genericnumber3, genericnumber4, genericnumber5, linenumber, orderid, origintypeid,
                   productdescription, productid, productname, sublinenumber, value, genericattribute27,
                   genericattribute28, genericattribute29, genericattribute30, genericattribute22, genericattribute23,
                   genericattribute24, genericattribute31, genericboolean1, genericboolean2
            FROM ZSALESTRANSACTION_LIST
            ORDER BY salestransactionseq DESC
            LIMIT ? OFFSET ?
        """
        cursor.execute(data_query, (per_page, (page - 1) * per_page))

    data = cursor.fetchall()
    cursor.close()
    conn.close()

    total_pages = math.ceil(total / per_page)

    # If download is requested, fetch and export all data to Excel
    if download:
        try:
            return download_sales_data(search_query, search_field, start_date, end_date)
        except ValueError as ve:
            flash(str(ve), 'danger')
            return redirect(url_for('indexhistorySalesTransactions', search=search_query, field=search_field))
        except Exception as e:
            app.logger.error(f"Error exporting data: {e}")
            flash("An unexpected error occurred. Please try again later.", 'danger')
            return redirect(url_for('indexhistorySalesTransactions', search=search_query, field=search_field))

    return render_template(
        'indexhistorySalesTransactions.html',
        data=data,
        page=page,
        total_pages=total_pages,
        search_query=search_query,
        search_field=search_field
    )


# Function to fetch all data from the database for exporting to Excel
def fetch_sales_data(search_query=None, search_field=None, start_date=None, end_date=None):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Base query to select data from ZSALESTRANSACTION_LIST
    query = """
        SELECT salestransactionseq, comments, compensationdate, eventtype, genericattribute1, genericattribute2,
               genericattribute3, genericattribute6, genericattribute12, genericnumber1, genericnumber2,
               genericnumber3, genericnumber4, genericnumber5, linenumber, orderid, origintypeid,
               productdescription, productid, productname, sublinenumber, value, genericattribute27,
               genericattribute28, genericattribute29, genericattribute30, genericattribute22, genericattribute23,
               genericattribute24, genericattribute31, genericboolean1, genericboolean2
        FROM ZSALESTRANSACTION_LIST
    """

    params = []
    conditions = []

    # Add search query if provided
    if search_query and search_field:
        conditions.append(f"{search_field} LIKE ?")
        params.append(f"%{search_query}%")

    # Add date range filter if start_date and end_date are provided
    if start_date and end_date:
        conditions.append("compensationdate BETWEEN ? AND ?")
        params.extend([start_date, end_date])

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    cursor.execute(query, params)
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return data


# Function to stream Excel data as it's generated for download
def generate_excel_stream(data, headers, max_rows_per_sheet=1000000):
    # Create an in-memory Excel file
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})

    total_rows = len(data)
    sheet_number = 1

    for start_idx in range(0, total_rows, max_rows_per_sheet):
        end_idx = min(start_idx + max_rows_per_sheet, total_rows)
        sheet_data = data[start_idx:end_idx]

        # Add new sheet
        worksheet = workbook.add_worksheet(f"Sheet{sheet_number}")
        sheet_number += 1

        # Write headers to the first row of the sheet
        worksheet.write_row(0, 0, headers)

        # Write data row by row starting from the second row (index 1)
        for row_idx, row_data in enumerate(sheet_data, start=1):
            worksheet.write_row(row_idx, 0, row_data)

    workbook.close()
    output.seek(0)

    return output.getvalue()


@app.route('/download_sales_data', methods=['GET'])
def download_sales_data(search_query=None, search_field=None, start_date=None, end_date=None):
    # Fetch data from the database
    data = fetch_sales_data(search_query, search_field, start_date, end_date)
    
    headers = [
        "salestransactionseq", "comments", "compensationdate", "eventtype", "genericattribute1", "genericattribute2",
        "genericattribute3", "genericattribute6", "genericattribute12", "genericnumber1", "genericnumber2",
        "genericnumber3", "genericnumber4", "genericnumber5", "linenumber", "orderid", "origintypeid",
        "productdescription", "productid", "productname", "sublinenumber", "value", "genericattribute27",
        "genericattribute28", "genericattribute29", "genericattribute30", "genericattribute22", "genericattribute23",
        "genericattribute24", "genericattribute31", "genericboolean1", "genericboolean2"
    ]

    # Stream Excel file as it's generated
    def generate():
        yield generate_excel_stream(data, headers)

    return Response(generate(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=sales_history.xlsx"})




@app.route('/viewHistory/<int:salestransactionseq>', methods=['GET'])
def viewHistory(salestransactionseq):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Fetch the history's data
    cursor.execute("""
        SELECT salestransactionseq, comments, compensationdate, eventtype, genericattribute1, genericattribute2, genericattribute3, genericattribute6, genericattribute12, genericnumber1, genericnumber2, genericnumber3, genericnumber4, genericnumber5, linenumber, orderid, origintypeid, productdescription, productid, productname, sublinenumber, value, genericattribute27, genericattribute28, genericattribute29, genericattribute30, genericattribute22, genericattribute23, genericattribute24, genericattribute31, genericboolean1, genericboolean2
        FROM ZSALESTRANSACTION_LIST
        WHERE salestransactionseq = ?
    """, (salestransactionseq,))
    history_data = cursor.fetchone()
    cursor.close()
    conn.close()

    if history_data is None:
        flash('Retailer not found.', 'error')
        return redirect(url_for('indexhistorySalesTransactions'))

    return render_template('viewHistory.html', history=history_data)



# #ARCHIVED CAMPAIGN ERROR Table 

@app.route('/indexArchived')
def indexArchived():
    search_query = request.args.get('search')
    search_field = request.args.get('field', 'salestransactionseq')  # Default search field
    page = request.args.get('page', 1, type=int)
    per_page = 8

    conn = get_db_connection()
    cursor = conn.cursor()

    valid_fields = [
        'salestransactionseq', 'comments', 'compensationdate', 'eventtype', 'genericattribute1', 'genericattribute2', 'genericattribute3', 
        'genericattribute6', 'genericattribute12', 'genericnumber1', 'genericnumber2', 'genericnumber3', 'genericnumber4', 'genericnumber5', 
        'linenumber', 'orderid', 'origintypeid', 'productdescription', 'productid', 'productname', 'sublinenumber', 'value', 
        'genericattribute27', 'genericattribute28', 'genericattribute29', 'genericattribute30', 'genericattribute22', 
        'genericattribute23', 'genericattribute24', 'genericattribute31', 'genericboolean1', 'genericboolean2','error'
    ]

    if search_query and search_field in valid_fields:
        query = f"SELECT COUNT(*) FROM ZARCHIVEDCAMPAIGN_ERROR WHERE {search_field} LIKE ?"
        cursor.execute(query, (f"%{search_query}%",))
        total = cursor.fetchone()[0]

        query = f"""
        SELECT salestransactionseq, comments, compensationdate, eventtype, genericattribute1, genericattribute2, genericattribute3, 
        genericattribute6, genericattribute12, genericnumber1, genericnumber2, genericnumber3, genericnumber4, genericnumber5, 
        linenumber, orderid, origintypeid, productdescription, productid, productname, sublinenumber, value, 
        genericattribute27, genericattribute28, genericattribute29, genericattribute30, genericattribute22, 
        genericattribute23, genericattribute24, genericattribute31, genericboolean1, genericboolean2, error
        FROM ZARCHIVEDCAMPAIGN_ERROR
        WHERE {search_field} LIKE ?
        ORDER BY CAST(salestransactionseq AS INTEGER) DESC
        LIMIT ? OFFSET ?
        """
        cursor.execute(query, (f"%{search_query}%", per_page, (page - 1) * per_page))
    else:
        query = "SELECT COUNT(*) FROM ZARCHIVEDCAMPAIGN_ERROR"
        cursor.execute(query)
        total = cursor.fetchone()[0]

        query = """
        SELECT salestransactionseq, comments, compensationdate, eventtype, genericattribute1, genericattribute2, genericattribute3, 
        genericattribute6, genericattribute12, genericnumber1, genericnumber2, genericnumber3, genericnumber4, genericnumber5, 
        linenumber, orderid, origintypeid, productdescription, productid, productname, sublinenumber, value, 
        genericattribute27, genericattribute28, genericattribute29, genericattribute30, genericattribute22, 
        genericattribute23, genericattribute24, genericattribute31, genericboolean1, genericboolean2, error
        FROM ZARCHIVEDCAMPAIGN_ERROR
        ORDER BY CAST(salestransactionseq AS INTEGER) DESC
        LIMIT ? OFFSET ?
        """
        cursor.execute(query, (per_page, (page - 1) * per_page))

    data = cursor.fetchall()
    cursor.close()
    conn.close()

    total_pages = math.ceil(total / per_page)

    if 'download' in request.args:
        return export_archived_excel(search_query, search_field)

    return render_template('indexArchived.html', data=data, page=page, total_pages=total_pages, search_query=search_query, search_field=search_field)

def export_archived_excel(search_query=None, search_field=None):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Define the batch size for fetching rows
    batch_size = 10000
    offset = 0
    all_rows = []

    while True:
        if search_query:
            query = f"""
            SELECT salestransactionseq, comments, compensationdate, eventtype, genericattribute1, genericattribute2,
                   genericattribute3, genericattribute6, genericattribute12, genericnumber1, genericnumber2,
                   genericnumber3, genericnumber4, genericnumber5, linenumber, orderid, origintypeid, 
                   productdescription, productid, productname, sublinenumber, value, genericattribute27, 
                   genericattribute28, genericattribute29, genericattribute30, genericattribute22, 
                   genericattribute23, genericattribute24, genericattribute31, genericboolean1, genericboolean2, error
            FROM ZARCHIVEDCAMPAIGN_ERROR
            WHERE {search_field} LIKE ?
            ORDER BY CAST(salestransactionseq AS INTEGER) DESC
            LIMIT ? OFFSET ?
            """
            params = (f"%{search_query}%", batch_size, offset)
        else:
            query = f"""
            SELECT salestransactionseq, comments, compensationdate, eventtype, genericattribute1, genericattribute2,
                   genericattribute3, genericattribute6, genericattribute12, genericnumber1, genericnumber2,
                   genericnumber3, genericnumber4, genericnumber5, linenumber, orderid, origintypeid, 
                   productdescription, productid, productname, sublinenumber, value, genericattribute27, 
                   genericattribute28, genericattribute29, genericattribute30, genericattribute22, 
                   genericattribute23, genericattribute24, genericattribute31, genericboolean1, genericboolean2, error
            FROM ZARCHIVEDCAMPAIGN_ERROR
            ORDER BY CAST(salestransactionseq AS INTEGER) DESC
            LIMIT ? OFFSET ?
            """
            params = (batch_size, offset)

        cursor.execute(query, params)
        rows = cursor.fetchall()

        if not rows:
            break

        all_rows.extend(rows)
        offset += batch_size

    conn.close()

    # Define column names based on the SELECT statement
    columns = [
        'SalesTransactionSeq', 'Comments', 'CompensationDate', 'EventType', 'RetailerMSISDN', 'ConsumerMSISDN', 
        'FranchiseID', 'CorelationID', 'TransactionID', 'Commission', 'GSTResult%', 'STWResult%', 'WHTResult%', 
        'NetCommission', 'LineNumber', 'OrderID', 'OriginTypeID', 'ProductDescription', 'ProductID', 
        'ProductName', 'SubLineNumber', 'Value', 'Status', 'SalesTaxWithheld%', 'GST%', 'WithHoldingTax%', 
        'City', 'Region', 'RSOCODE', 'AggregatedTaxRate%', 'TaxExemptedFlag', 'RetailerStatus', 'Error'
    ]

    # Create a DataFrame from the fetched rows
    df = pd.DataFrame(all_rows, columns=columns)

    # Create an in-memory Excel file
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='ArchivedHistory')

    output.seek(0)

    # Send the Excel file as a response
    return send_file(output, as_attachment=True, download_name='archived_history.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/viewArchive/<int:salestransactionseq>', methods=['GET'])
def viewArchive(salestransactionseq):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Fetch the product's data
    cursor.execute("""
        SELECT salestransactionseq, comments, compensationdate, eventtype, genericattribute1, genericattribute2, genericattribute3, genericattribute6, genericattribute12, genericnumber1, genericnumber2, genericnumber3, genericnumber4, genericnumber5, linenumber, orderid, origintypeid, productdescription, productid, productname, sublinenumber, value, genericattribute27, genericattribute28, genericattribute29, genericattribute30, genericattribute22, genericattribute23, genericattribute24, genericattribute31, genericboolean1, genericboolean2, error
        FROM ZARCHIVEDCAMPAIGN_ERROR 
        WHERE salestransactionseq = ?
    """, (salestransactionseq,))
    error_data = cursor.fetchone()
    cursor.close()
    conn.close()

    if error_data is None:
        flash('Retailer not found.', 'error')
        return redirect(url_for('indexTransactionsError'))

    return render_template('viewArchive.html', error=error_data)


# Sales Transaction Error Table

@app.route('/TransactionsError')
def indexTransactionsError():
    search_query = request.args.get('search')
    search_field = request.args.get('field', 'salestransactionseq')  # Default search field
    page = request.args.get('page', 1, type=int)
    per_page = 8

    # Get download parameters
    download = request.args.get('download', False)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    conn = get_db_connection()
    cursor = conn.cursor()

    # List of valid search fields
    valid_fields = [
        'salestransactionseq', 'comments', 'compensationdate', 'error_msg', 'eventtype', 'genericattribute1', 'genericattribute2',
        'genericattribute3', 'genericattribute6', 'genericattribute12', 'genericnumber1', 'genericnumber2', 'genericnumber3',
        'genericnumber4', 'genericnumber5', 'linenumber', 'orderid', 'origintypeid', 'productdescription', 'productid', 'productname',
        'sublinenumber', 'value', 'genericattribute27', 'genericattribute28', 'genericattribute29', 'genericattribute30', 'channel',
        'genericattribute22', 'genericattribute23', 'genericattribute24', 'genericattribute31', 'genericboolean1', 'genericboolean2'
    ]

    # Query for counting total records based on search criteria
    if search_query and search_field in valid_fields:
        count_query = f"SELECT COUNT(*) FROM zsalestransaction_error WHERE {search_field} LIKE ?"
        params = (f"%{search_query}%",)
        cursor.execute(count_query, params)
        total = cursor.fetchone()[0]

        # Data query to retrieve paginated results based on search query
        data_query = f"""
            SELECT salestransactionseq, comments, compensationdate, error_msg, eventtype, genericattribute1, genericattribute2,
                   genericattribute3, genericattribute6, genericattribute12, genericnumber1, genericnumber2, genericnumber3,
                   genericnumber4, genericnumber5, linenumber, orderid, origintypeid, productdescription, productid, productname,
                   sublinenumber, value, genericattribute27, genericattribute28, genericattribute29, genericattribute30, channel,
                   genericattribute22, genericattribute23, genericattribute24, genericattribute31, genericboolean1, genericboolean2
            FROM zsalestransaction_error
            WHERE {search_field} LIKE ?
            ORDER BY CAST(salestransactionseq AS INTEGER) DESC
            LIMIT ? OFFSET ?
        """
        cursor.execute(data_query, (f"%{search_query}%", per_page, (page - 1) * per_page))
    else:
        # Query for counting total records without search query
        count_query = "SELECT COUNT(*) FROM zsalestransaction_error"
        cursor.execute(count_query)
        total = cursor.fetchone()[0]

        # Data query to retrieve paginated results
        data_query = """
            SELECT salestransactionseq, comments, compensationdate, error_msg, eventtype, genericattribute1, genericattribute2,
                   genericattribute3, genericattribute6, genericattribute12, genericnumber1, genericnumber2, genericnumber3,
                   genericnumber4, genericnumber5, linenumber, orderid, origintypeid, productdescription, productid, productname,
                   sublinenumber, value, genericattribute27, genericattribute28, genericattribute29, genericattribute30, channel,
                   genericattribute22, genericattribute23, genericattribute24, genericattribute31, genericboolean1, genericboolean2
            FROM zsalestransaction_error
            ORDER BY CAST(salestransactionseq AS INTEGER) DESC
            LIMIT ? OFFSET ?
        """
        cursor.execute(data_query, (per_page, (page - 1) * per_page))

    data = cursor.fetchall()
    cursor.close()
    conn.close()

    total_pages = math.ceil(total / per_page)

    # If download is requested, fetch and export all data to Excel
    if download:
        try:
            return export_sales_transactions_error_to_excel(search_query, search_field, start_date, end_date)
        except ValueError as ve:
            flash(str(ve), 'danger')
            return redirect(url_for('indexTransactionsError', search=search_query, field=search_field))
        except Exception as e:
            app.logger.error(f"Error exporting data: {e}")
            flash("An unexpected error occurred. Please try again later.", 'danger')
            return redirect(url_for('indexTransactionsError', search=search_query, field=search_field))

    return render_template(
        'indexTransactionsError.html',
        data=data,
        page=page,
        total_pages=total_pages,
        search_query=search_query,
        search_field=search_field
    )


# Function to fetch all data from the database for exporting to Excel
def fetch_sales_transactions_error_data(search_query=None, search_field=None, start_date=None, end_date=None):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Base query to select data from zsalestransaction_error
    query = """
        SELECT salestransactionseq, comments, compensationdate, error_msg, eventtype, genericattribute1, genericattribute2,
               genericattribute3, genericattribute6, genericattribute12, genericnumber1, genericnumber2, genericnumber3,
               genericnumber4, genericnumber5, linenumber, orderid, origintypeid, productdescription, productid, productname,
               sublinenumber, value, genericattribute27, genericattribute28, genericattribute29, genericattribute30, channel,
               genericattribute22, genericattribute23, genericattribute24, genericattribute31, genericboolean1, genericboolean2
        FROM zsalestransaction_error
    """

    params = []
    conditions = []

    # Add search query if provided
    if search_query and search_field:
        conditions.append(f"{search_field} LIKE ?")
        params.append(f"%{search_query}%")

    # Add date range filter if start_date and end_date are provided
    if start_date and end_date:
        conditions.append("compensationdate BETWEEN ? AND ?")
        params.extend([start_date, end_date])

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    cursor.execute(query, params)
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return data


# Function to stream Excel data as it's generated for download
def generate_excel_stream_for_error_data(data, headers, max_rows_per_sheet=1000000):
    # Create an in-memory Excel file
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})

    total_rows = len(data)
    sheet_number = 1

    for start_idx in range(0, total_rows, max_rows_per_sheet):
        end_idx = min(start_idx + max_rows_per_sheet, total_rows)
        sheet_data = data[start_idx:end_idx]

        # Add new sheet
        worksheet = workbook.add_worksheet(f"Sheet{sheet_number}")
        sheet_number += 1

        # Write headers to the first row of the sheet
        worksheet.write_row(0, 0, headers)

        # Write data row by row starting from the second row (index 1)
        for row_idx, row_data in enumerate(sheet_data, start=1):
            worksheet.write_row(row_idx, 0, row_data)

    workbook.close()
    output.seek(0)

    return output.getvalue()


@app.route('/download_sales_transactions_error_data', methods=['GET'])
def export_sales_transactions_error_to_excel(search_query=None, search_field=None, start_date=None, end_date=None):
    # Fetch data from the database
    data = fetch_sales_transactions_error_data(search_query, search_field, start_date, end_date)
    
    headers = [
        "salestransactionseq", "comments", "compensationdate", "error_msg", "eventtype", "genericattribute1", "genericattribute2",
        "genericattribute3", "genericattribute6", "genericattribute12", "genericnumber1", "genericnumber2", "genericnumber3",
        "genericnumber4", "genericnumber5", "linenumber", "orderid", "origintypeid", "productdescription", "productid", "productname",
        "sublinenumber", "value", "genericattribute27", "genericattribute28", "genericattribute29", "genericattribute30", "channel",
        "genericattribute22", "genericattribute23", "genericattribute24", "genericattribute31", "genericboolean1", "genericboolean2"
    ]

    # Stream Excel file as it's generated
    def generate():
        yield generate_excel_stream_for_error_data(data, headers)

    return Response(generate(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=error_transactions.xlsx"})



@app.route('/viewErrors/<int:salestransactionseq>', methods=['GET'])
def viewErrors(salestransactionseq):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Fetch the product's data
    cursor.execute("""
        SELECT salestransactionseq, comments, compensationdate, error_msg, eventtype, genericattribute1, genericattribute2, genericattribute3, genericattribute6, genericattribute12, genericnumber1, genericnumber2, genericnumber3, genericnumber4, genericnumber5, linenumber, orderid, origintypeid, productdescription, productid, productname, sublinenumber, value, genericattribute27, genericattribute28, genericattribute29, genericattribute30, channel, genericattribute22, genericattribute23, genericattribute24, genericattribute31, genericboolean1, genericboolean2
        FROM zsalestransaction_error 
        WHERE salestransactionseq = ?
    """, (salestransactionseq,))
    error_data = cursor.fetchone()
    cursor.close()
    conn.close()

    if error_data is None:
        flash('Retailer not found.', 'error')
        return redirect(url_for('indexTransactionsError'))

    return render_template('viewErrors.html', error=error_data)


@app.route('/clean_data', methods=['POST'])
def clean_data():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Execute the stored procedure
        cursor.execute("CALL MOVE_ERROR_RECORDS_TO_ARCHIVE()")
        conn.commit()

        cursor.close()
        conn.close()

        flash('Data is cleaned and moved to the archive table successfully.', 'success')
    except Exception as e:
        error_message = str(e)
        
        # Check for specific error patterns
        if 'no data found' in error_message:
            friendly_message = "No data was found to move to the archive."
        else:
            # General fallback error message
            friendly_message = "An unexpected error occurred. Please try again later."

        flash(friendly_message, 'error')

    return redirect(url_for('indexTransactionsError'))

#DISPLAY TABLE ONLY

@app.route('/DisplayTransactionsError')
def DisplayTransactionsError():
    search_query = request.args.get('search')
    search_field = request.args.get('field', 'salestransactionseq')  # Default search field
    page = request.args.get('page', 1, type=int)
    per_page = 8

    conn = get_db_connection()
    cursor = conn.cursor()

    valid_fields = ['salestransactionseq', 'comments', 'compensationdate', 'error_msg', 'eventtype', 'genericattribute1', 'genericattribute2', 'genericattribute3', 'genericattribute6', 'genericattribute12', 'genericnumber1', 'genericnumber2', 'genericnumber3', 'genericnumber4', 'genericnumber5', 'linenumber', 'orderid', 'origintypeid', 'productdescription', 'productid', 'productname', 'sublinenumber', 'value', 'genericattribute27', 'genericattribute28', 'genericattribute29', 'genericattribute30', 'channel']
    
    if search_query and search_field in valid_fields:
        # Count total records for pagination
        query = f"SELECT COUNT(*) FROM zsalestransaction_error WHERE {search_field} LIKE ?"
        cursor.execute(query, (f"%{search_query}%",))
        total = cursor.fetchone()[0]

        # Fetch paginated records
        query = f"""
        SELECT salestransactionseq, comments, compensationdate, error_msg, eventtype, genericattribute1, genericattribute2, genericattribute3, genericattribute6, genericattribute12, genericnumber1, genericnumber2, genericnumber3, genericnumber4, genericnumber5, linenumber, orderid, origintypeid, productdescription, productid, productname, sublinenumber, value, genericattribute27, genericattribute28, genericattribute29, genericattribute30, channel
        FROM zsalestransaction_error
        WHERE {search_field} LIKE ?
        ORDER BY salestransactionseq ASC
        LIMIT ? OFFSET ?
        """
        cursor.execute(query, (f"%{search_query}%", per_page, (page - 1) * per_page))
    else:
        # Count total records for pagination
        query = "SELECT COUNT(*) FROM zsalestransaction_error"
        cursor.execute(query)
        total = cursor.fetchone()[0]

        # Fetch paginated records
        query = """
        SELECT salestransactionseq, comments, compensationdate, error_msg, eventtype, genericattribute1, genericattribute2, genericattribute3, genericattribute6, genericattribute12, genericnumber1, genericnumber2, genericnumber3, genericnumber4, genericnumber5, linenumber, orderid, origintypeid, productdescription, productid, productname, sublinenumber, value, genericattribute27, genericattribute28, genericattribute29, genericattribute30
        FROM zsalestransaction_error
        ORDER BY salestransactionseq ASC
        LIMIT ? OFFSET ?
        """
        cursor.execute(query, (per_page, (page - 1) * per_page))

    data = cursor.fetchall()
    cursor.close()
    conn.close()

    total_pages = math.ceil(total / per_page)
    return render_template('DisplayTransactionsError.html', data=data, page=page, total_pages=total_pages, search_query=search_query, search_field=search_field)


@app.route('/editTransactions/<int:salestransactionseq>', methods=['GET', 'POST'])
def editTransactionsError(salestransactionseq):
    conn = get_db_connection()
    cursor = conn.cursor()
    if request.method == 'POST':
        # Fetch form data and handle potential missing data
        try:
            new_comments = request.form['comments']
            new_compensationdate = request.form['compensationdate']
            new_error_msg = request.form['error_msg']
            new_eventtype = request.form['eventtype']
            new_genericattribute1 = request.form['genericattribute1']
            new_genericattribute2 = request.form['genericattribute2']
            new_genericattribute3 = request.form['genericattribute3']
            new_genericattribute6 = request.form['genericattribute6']
            new_genericattribute12 = request.form['genericattribute12']
            new_genericnumber1 = request.form['genericnumber1']
            new_genericnumber2 = request.form['genericnumber2']
            new_genericnumber3 = request.form['genericnumber3']
            new_genericnumber4 = request.form['genericnumber4']
            new_genericnumber5 = request.form['genericnumber5']
            new_linenumber = request.form['linenumber']
            new_orderid = request.form['orderid']
            new_origintypeid = request.form['origintypeid']
            new_productdescription = request.form['productdescription']
            new_productid = request.form['productid']
            new_productname = request.form['productname']
            new_sublinenumber = request.form['sublinenumber']
            new_value = request.form['value']
            new_genericattribute27 = request.form['genericattribute27']
            new_genericattribute28 = request.form['genericattribute28']
            new_genericattribute29 = request.form['genericattribute29']
            new_genericattribute30 = request.form['genericattribute30']
       

            # Convert to integers where necessary
            new_linenumber = int(new_linenumber)
            new_sublinenumber = int(new_sublinenumber)
        except ValueError as ve:
            flash(f'Error with data: {ve}')
            return redirect(url_for('editTransactionsError', salestransactionseq=salestransactionseq))
        except KeyError as ke:
            flash(f'Missing form field: {ke}')
            return redirect(url_for('editTransactionsError', salestransactionseq=salestransactionseq))

        # Update the record
        try:
            cursor.execute("""
            UPDATE zsalestransaction_error
            SET comments = ?, compensationdate = ?, error_msg = ?, eventtype = ?, 
                genericattribute1 = ?, genericattribute2 = ?, genericattribute3 = ?, genericattribute6 = ?, genericattribute12 = ?,
                genericnumber1 = ?, genericnumber2 = ?, genericnumber3 = ?, genericnumber4 = ?, genericnumber5 = ?, linenumber = ?, orderid = ?, origintypeid = ?, productdescription = ?, 
                productid = ?, productname = ?, sublinenumber = ?, value = ?, 
                genericattribute27 = ?, genericattribute28 = ?, genericattribute29 = ?, genericattribute30 = ?
            WHERE salestransactionseq = ?
            """, (new_comments, new_compensationdate, new_error_msg, new_eventtype, new_genericattribute1, 
                  new_genericattribute2, new_genericattribute3, new_genericattribute6, new_genericattribute12, new_genericnumber1, 
                  new_genericnumber2, new_genericnumber3, new_genericnumber4, new_genericnumber5, 
                  new_linenumber, new_orderid, new_origintypeid, new_productdescription, 
                  new_productid, new_productname, new_sublinenumber, new_value, new_genericattribute27, new_genericattribute28, new_genericattribute29, new_genericattribute30, salestransactionseq))
            conn.commit()
            flash('Entry successfully updated.')
        except dbapi.Error as e:
            flash(f'Database error: {e}')
        finally:
            cursor.close()
            conn.close()

        return redirect(url_for('indexTransactionsError'))
    
    cursor.execute("""
    SELECT  salestransactionseq, comments, compensationdate, error_msg, eventtype, genericattribute1, genericattribute2, genericattribute3, genericattribute6, genericattribute12, genericnumber1, genericnumber2, genericnumber3, genericnumber4, genericnumber5, linenumber, orderid, origintypeid, productdescription, productid, productname, sublinenumber, value, genericattribute27, genericattribute28, genericattribute29, genericattribute30
    FROM zsalestransaction_error 
    WHERE salestransactionseq = ?
    """, (salestransactionseq,))
    entry = cursor.fetchone()
    cursor.close()
    conn.close()

    if not entry:
        flash('Entry not found.')
        return redirect(url_for('indexTransactionsError'))

    return render_template('editTransactionsError.html', entry=entry)


@app.route('/deleteTransactions/<int:salestransactionseq>', methods=['POST'])
def deleteTransactionsError(salestransactionseq):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM zsalestransaction_error WHERE salestransactionseq = ?", (salestransactionseq,))
    conn.commit()
    cursor.close()
    conn.close()
    flash('Entry successfully deleted.')
    return redirect(url_for('indexTransactionsError'))


@app.route('/transactions/error/preprocess/<int:salestransactionseq>', methods=['POST'])
def preprocessTransaction(salestransactionseq):
    # Update the GENERICATTRIBUTE32 field to 1
    conn = get_db_connection()
    cursor = conn.cursor()
    query = "UPDATE ZSALESTRANSACTION_ERROR SET GENERICATTRIBUTE32 = 1 WHERE salestransactionseq = ?"
    cursor.execute(query, (salestransactionseq,))
    conn.commit()
    cursor.close()
    conn.close()
    
    # flash("Transaction preprocessed successfully", "success")
    return redirect(url_for('indexTransactionsError'))


@app.route('/transactions/error/bulk_reprocess', methods=['POST'])
def bulk_reprocess_transactions():
    conn = get_db_connection()
    cursor = conn.cursor()

    query = "UPDATE ZSALESTRANSACTION_ERROR SET GENERICATTRIBUTE32 = 1"
    
    try:
        cursor.execute(query)
        conn.commit()
        flash("All transactions reprocessed successfully.", "success")
    except dbapi.Error as e:
        flash(f"Database error during bulk reprocess: {e}", "danger")
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('indexTransactionsError'))


@app.route('/transactions/error/bulk_update', methods=['POST'])
def bulk_update_transactions():
    field_to_update = request.form['field_to_update']
    new_value = request.form['new_value']

    # Sanitize the input to prevent SQL injection (field name must be validated)
    valid_fields = ['comments', 'compensationdate', 'error_msg', 'eventtype', 'genericattribute1', 
                    'genericattribute2', 'genericattribute3', 'genericattribute6', 'genericattribute12', 
                    'genericnumber1', 'genericnumber2', 'genericnumber3', 'genericnumber4', 
                    'genericnumber5', 'linenumber', 'orderid', 'origintypeid', 'productdescription', 
                    'productid', 'productname', 'sublinenumber', 'value', 
                    'genericattribute27', 'genericattribute28', 'genericattribute29', 'genericattribute30']

    if field_to_update not in valid_fields:
        flash('Invalid field selected for update.', 'danger')
        return redirect(url_for('indexTransactionsError'))

    # Prepare the SQL query
    conn = get_db_connection()
    cursor = conn.cursor()
    
    query = f"UPDATE zsalestransaction_error SET {field_to_update} = ?"
    
    try:
        cursor.execute(query, (new_value,))
        conn.commit()
        flash(f'All records updated successfully for {field_to_update}.', 'success')
    except dbapi.Error as e:
        flash(f'Database error during bulk update: {e}', 'danger')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('indexTransactionsError'))


# ZMDLT TABLE 


@app.route('/Zmdlt', methods=['GET', 'POST'])
def indexZmdlt():
    if request.method == 'POST':
        excel_file = request.files.get('file')
        if excel_file and excel_file.filename.endswith('.xlsx'):
            try:
                # Read the Excel file
                df = pd.read_excel(excel_file)
                print(df)  # Debug: Show the raw Excel content

                # Convert column names to upper case and strip spaces
                df.columns = df.columns.str.strip().str.upper()

                required_columns = [
                    'STARTDATE', 'ENDDATE', 'COMPAIGNNAME', 'FRANCHISEID',
                    'PRODUCTID', 'COMMISION'
                ]

                missing_columns = [col for col in required_columns if col not in df.columns]
                extra_columns = [col for col in df.columns if col not in required_columns]

                if missing_columns or extra_columns:
                    error_message = "Error: The following issues were found in the Excel file:\n"
                    if missing_columns:
                        error_message += f"Missing columns: {', '.join(missing_columns)}\n"
                    if extra_columns:
                        error_message += f"Extra columns: {', '.join(extra_columns)}\n"
                    flash(error_message.strip())
                    return redirect(url_for('indexZmdlt'))

                # Explicit date parsing with handling of future dates
                try:
                    df['STARTDATE'] = pd.to_datetime(df['STARTDATE'], format='%Y-%m-%d %H:%M:%S', errors='coerce')
                    df['ENDDATE'] = pd.to_datetime(df['ENDDATE'], format='%Y-%m-%d %H:%M:%S', errors='coerce')

                    # Identify rows where conversion failed (NaT) and log them
                    invalid_dates = df[df['STARTDATE'].isna() | df['ENDDATE'].isna()]
                    if not invalid_dates.empty:
                        print(f"Invalid date rows:\n{invalid_dates}")
                        
                        # Log or handle invalid rows (for example, by removing them)
                        df = df.drop(invalid_dates.index)

                except Exception as date_error:
                    flash(f"Error: One of the date columns (STARTDATE or ENDDATE) has invalid data. {str(date_error)}")
                    return redirect(url_for('indexZmdlt'))

                # Debug: Show the dataframe after date conversion
                print(df[['STARTDATE', 'ENDDATE']])

                # Remove rows where dates are invalid (NaT values) if any remain
                df.dropna(subset=['STARTDATE', 'ENDDATE'], inplace=True)

                # Rename columns as required for the database
                df.rename(columns={
                    'STARTDATE': 'EFFECTIVESTARTDATE',
                    'ENDDATE': 'EFFECTIVEENDDATE',
                    'COMPAIGNNAME': 'MDLTNAME',
                    'FRANCHISEID': 'DIM0',
                    'PRODUCTID': 'DIM1',
                    'COMMISION': 'VALUE'
                }, inplace=True)

                conn = get_db_connection()
                cursor = conn.cursor()

                # Get CITY and REGION information based on the uploaded Franchise IDs
                franchise_ids = tuple(df['DIM0'].unique())
                placeholder = ','.join('?' for _ in franchise_ids)
                cursor.execute(f"""
                    SELECT DISTINCT ZR.genericattribute5 AS FranchiseID, 
                                    ZR.genericattribute6 AS City, 
                                    ZR.genericattribute7 AS Region
                    FROM ZRETAILER ZR
                    WHERE ZR.genericattribute5 IN ({placeholder})
                """, franchise_ids)

                retailer_map = {
                    row[0]: {'city': row[1], 'region': row[2]}
                    for row in cursor.fetchall()
                }

                # Get valid product IDs
                cursor.execute("SELECT DISTINCT PRICE FROM ZPRODUCT")
                valid_product_ids = {row[0] for row in cursor.fetchall()}

                error_rows = []
                new_entries = []
                update_entries = []

                for index, row in df.iterrows():
                    row_errors = []

                    # Populate CITY and REGION based on DIM0 (FranchiseID)
                    franchise_info = retailer_map.get(row['DIM0'], {})
                    if franchise_info:
                        city = franchise_info['city']
                        region = franchise_info['region']
                    else:
                        row_errors.append(f"Invalid or missing FranchiseID (DIM0): {row['DIM0']}")

                    # Validate ProductID
                    if row['DIM1'] not in valid_product_ids:
                        row_errors.append(f"Invalid ProductID (DIM1): {row['DIM1']}")

                    # Check for missing values in required columns
                    for col in ['MDLTNAME', 'VALUE', 'DIM0', 'DIM1']:
                        if pd.isnull(row[col]):
                            row_errors.append(f"Missing value for {col}. Expected data type: {df[col].dtype}")

                    # Ensure no overlapping campaigns with different MDLTNAME within the same timeframe
                    cursor.execute("""
                        SELECT MDLTNAME, EFFECTIVESTARTDATE, EFFECTIVEENDDATE
                        FROM ZMDLT
                        WHERE DIM0 = ? AND DIM1 = ? 
                        AND (EFFECTIVESTARTDATE <= ? AND EFFECTIVEENDDATE >= ?)
                    """, (row['DIM0'], row['DIM1'], row['EFFECTIVEENDDATE'], row['EFFECTIVESTARTDATE']))

                    overlapping_campaigns = cursor.fetchall()

                    if overlapping_campaigns:
                        # Check if there's an overlapping campaign with a different name
                        for campaign in overlapping_campaigns:
                            if campaign[0] != row['MDLTNAME']:
                                row_errors.append(
                                    f"There is already a Campaign '{campaign[0]}' during the time period "
                                    f"{campaign[1].strftime('%Y-%m-%d')} to {campaign[2].strftime('%Y-%m-%d')}. "
                                    f"New campaign '{row['MDLTNAME']}' cannot be added in the same timeframe."
                                )

                    if row_errors:
                        error_rows.append(f"Row {index + 1}: " + "; ".join(row_errors))
                        continue

                    # Set eventtype to 93 as requested
                    eventtype = 93

                    # Check if entry exists for update or insert
                    cursor.execute("""
                        SELECT MDLTCELLSEQ, VALUE
                        FROM ZMDLT
                        WHERE DIM0 = ? AND DIM1 = ? AND MDLTNAME = ?
                    """, (row['DIM0'], row['DIM1'], row['MDLTNAME']))
                    record = cursor.fetchone()

                    modification_date = datetime.now()

                    if record:
                        existing_value = record[1]
                        if existing_value != row['VALUE']:
                            # Update if the commission value is different
                            update_entries.append((
                                row['EFFECTIVESTARTDATE'], row['EFFECTIVEENDDATE'], row['VALUE'],
                                'PKR', modification_date, city, region, eventtype, record[0]
                            ))
                    else:
                        new_entries.append((
                            row['EFFECTIVESTARTDATE'], row['EFFECTIVEENDDATE'], row['MDLTNAME'],
                            row['DIM0'], row['DIM1'], row['VALUE'], 'PKR',
                            modification_date, city, region, eventtype
                        ))

                # Batch update existing records
                if update_entries:
                    cursor.executemany("""
                        UPDATE ZMDLT
                        SET EFFECTIVESTARTDATE = ?, EFFECTIVEENDDATE = ?, VALUE = ?,
                            UNITTYPEFORVALUE = ?, MODIFICATIONDATE = ?, CITY = ?,
                            REGION = ?, eventtype = ?
                        WHERE MDLTCELLSEQ = ?
                    """, update_entries)

                # Batch insert new records
                if new_entries:
                    cursor.executemany("""
                        INSERT INTO ZMDLT (EFFECTIVESTARTDATE, EFFECTIVEENDDATE, MDLTNAME, DIM0, DIM1,
                                           VALUE, UNITTYPEFORVALUE, MODIFICATIONDATE, CITY, REGION, eventtype)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, new_entries)

                conn.commit()
                cursor.close()
                conn.close()

                if error_rows:
                    flash(f"Data processed with the following issues:\n" + "\n".join(error_rows))
                else:
                    flash('Excel file data successfully processed and added/updated in the database.')
            except Exception as e:
                flash(f'Error processing the file: {str(e)}')
        else:
            flash('Please upload a valid .xlsx file.')

        return redirect(url_for('indexZmdlt'))


    search_query = request.args.get('search')
    search_field = request.args.get('field', 'MDLTNAME').upper()
    page = request.args.get('page', 1, type=int)
    per_page = 8

    conn = get_db_connection()
    cursor = conn.cursor()

    if search_query:
        query = f"SELECT COUNT(*) FROM ZMDLT WHERE {search_field} LIKE ?"
        cursor.execute(query, (f"%{search_query}%",))
        total = cursor.fetchone()[0]

        query = f"""
        SELECT MDLTCELLSEQ, EFFECTIVESTARTDATE, EFFECTIVEENDDATE, MDLTNAME, VALUE, UNITTYPEFORVALUE, DIM0, DIM1, MODIFICATIONDATE, CITY, REGION, EVENTTYPE
        FROM ZMDLT
        WHERE {search_field} LIKE ?
        ORDER BY MDLTCELLSEQ DESC
        LIMIT ? OFFSET ?
        """
        cursor.execute(query, (f"%{search_query}%", per_page, (page - 1) * per_page))
    else:
        query = "SELECT COUNT(*) FROM ZMDLT"
        cursor.execute(query)
        total = cursor.fetchone()[0]

        query = """
        SELECT MDLTCELLSEQ, EFFECTIVESTARTDATE, EFFECTIVEENDDATE, MDLTNAME, VALUE, UNITTYPEFORVALUE, DIM0, DIM1, MODIFICATIONDATE, CITY, REGION, EVENTTYPE
        FROM ZMDLT
        ORDER BY MDLTCELLSEQ DESC
        LIMIT ? OFFSET ?
        """
        cursor.execute(query, (per_page, (page - 1) * per_page))

    data = cursor.fetchall()
    cursor.close()
    conn.close()

    total_pages = math.ceil(total / per_page)

    if 'download' in request.args:
        return export_zmdlt_to_excel(search_query, search_field)

    return render_template('indexZmdlt.html', data=data, page=page, total_pages=total_pages, search_query=search_query, search_field=search_field)


def export_zmdlt_to_excel(search_query=None, search_field=None):
    conn = get_db_connection()
    cursor = conn.cursor()

    batch_size = 10000
    offset = 0
    all_rows = []

    while True:
        if search_query:
            query = f"""
            SELECT EFFECTIVESTARTDATE, EFFECTIVEENDDATE, MDLTNAME, DIM0, DIM1, VALUE, UNITTYPEFORVALUE, 
                   MODIFICATIONDATE, CITY, REGION, EVENTTYPE
            FROM ZMDLT
            WHERE {search_field} LIKE ?
            ORDER BY MDLTCELLSEQ DESC
            LIMIT ? OFFSET ?
            """
            params = (f"%{search_query}%", batch_size, offset)
        else:
            query = f"""
            SELECT EFFECTIVESTARTDATE, EFFECTIVEENDDATE, MDLTNAME, DIM0, DIM1, VALUE, UNITTYPEFORVALUE, 
                   MODIFICATIONDATE, CITY, REGION, EVENTTYPE
            FROM ZMDLT
            ORDER BY MDLTCELLSEQ DESC
            LIMIT ? OFFSET ?
            """
            params = (batch_size, offset)

        cursor.execute(query, params)
        rows = cursor.fetchall()

        if not rows:
            break

        all_rows.extend(rows)
        offset += batch_size

    conn.close()

    df = pd.DataFrame(all_rows, columns=[
        'StartDate', 'EndDate', 'CompaignName', 'FranchiseID', 'ProductID',
        'Commision', 'UnitType', 'ModificationDate', 'City', 'Region', 'EventType'
    ])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='ZMDLT')

    output.seek(0)

    return send_file(output, as_attachment=True, download_name='ZmdltData.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/addZmdltEntry', methods=['GET', 'POST'])
def addZmdltEntry():
    conn = get_db_connection()
    cursor = conn.cursor()

    # Fetch franchise and product data for form options
    cursor.execute("SELECT DISTINCT genericattribute5 FROM ZRETAILER")
    franchises = cursor.fetchall()

    cursor.execute("SELECT DISTINCT PRICE FROM ZPRODUCT")
    products = cursor.fetchall()

    if request.method == 'POST':
        modificationdate = datetime.now()
        effectivestartdate = request.form['effectivestartdate']
        effectiveenddate = request.form['effectiveenddate']
        mdltname = request.form['mdltname']
        dim0 = request.form['dim0']
        dim1 = request.form['dim1']
        value = request.form['value']

        # Validate date format
        try:
            effectivestartdate = datetime.strptime(effectivestartdate, '%Y-%m-%d')
            effectiveenddate = datetime.strptime(effectiveenddate, '%Y-%m-%d')
        except ValueError:
            flash("Invalid date format. Please use YYYY-MM-DD.")
            return redirect(url_for('addZmdltEntry'))

        eventtype = 93  # Hardcoded event type

        # Fetch CITY and REGION based on FranchiseID
        cursor.execute("""
            SELECT genericattribute6, genericattribute7 
            FROM ZRETAILER 
            WHERE genericattribute5 = ?
        """, (dim0,))
        result = cursor.fetchone()

        if result:
            city, region = result
            city = city if city else "None"
            region = region if region else "None"
        else:
            flash("Invalid FranchiseID.")
            return redirect(url_for('addZmdltEntry'))

        # Fetch ALL existing campaigns to check for date overlaps with any campaign in the table
        cursor.execute("""
            SELECT EFFECTIVESTARTDATE, EFFECTIVEENDDATE, MDLTNAME 
            FROM ZMDLT 
            WHERE (EFFECTIVESTARTDATE <= ? AND EFFECTIVEENDDATE >= ?)
        """, (effectiveenddate, effectivestartdate))
        existing_campaigns = cursor.fetchall()

        # Check for date overlaps with different MDLTNAME
        for existing_start, existing_end, existing_mdltname in existing_campaigns:
            # If the names are different and the dates overlap, generate an error
            if existing_mdltname != mdltname:
                flash(f"There is already a Campaign '{existing_mdltname}' during this time period "
                      f"{existing_start.date()} to {existing_end.date()}. "
                      "Please change the Effective Start Date or End Date.")
                cursor.close()
                conn.close()
                return redirect(url_for('indexZmdlt'))

        try:
            # Check if the entry already exists for update or insert based on the same campaign name and dimensions
            cursor.execute("""
                SELECT COUNT(*) FROM ZMDLT 
                WHERE mdltname = ? AND dim0 = ? AND dim1 = ?
            """, (mdltname, dim0, dim1))
            count = cursor.fetchone()[0]

            if count > 0:
                # Update existing entry
                cursor.execute("""
                    UPDATE ZMDLT 
                    SET effectivestartdate = ?, effectiveenddate = ?, value = ?, 
                        unittypeforvalue = ?, modificationdate = ?, city = ?, 
                        region = ?, eventtype = ? 
                    WHERE mdltname = ? AND dim0 = ? AND dim1 = ?
                """, (effectivestartdate, effectiveenddate, value, 'PKR',
                      modificationdate, city, region, eventtype, mdltname, dim0, dim1))
                flash('ZMDLT entry updated successfully.')
            else:
                # Insert new entry
                cursor.execute("""
                    INSERT INTO ZMDLT (effectivestartdate, effectiveenddate, mdltname, value,
                                        dim0, dim1, modificationdate, city, region, eventtype, unittypeforvalue)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'PKR')
                """, (effectivestartdate, effectiveenddate, mdltname, value,
                      dim0, dim1, modificationdate, city, region, eventtype))
                flash('New ZMDLT entry added successfully.')

            conn.commit()
        except Exception as e:
            flash(f'An error occurred: {str(e)}')
        finally:
            cursor.close()
            conn.close()

        return redirect(url_for('indexZmdlt'))

    cursor.close()
    conn.close()

    return render_template('addZmdltEntry.html', franchises=franchises, products=products)


@app.route('/editZmdlt/<int:mdltcellseq>', methods=['GET', 'POST'])
def editZmdlt(mdltcellseq):
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT DISTINCT genericattribute5 FROM ZRETAILER")
    franchises = cursor.fetchall()

    cursor.execute("SELECT DISTINCT PRICE FROM ZPRODUCT")
    products = cursor.fetchall()

    if request.method == 'POST':
        new_modificationdate = datetime.now()
        new_effectivestartdate = request.form['effectivestartdate']
        new_effectiveenddate = request.form['effectiveenddate']
        new_mdltname = request.form['mdltname']
        new_value = request.form['value']
        new_dim0 = request.form['dim0']
        new_dim1 = request.form['dim1']

        # Set eventtype to 93 as requested
        new_eventtype = 93

        # Fetch CITY and REGION based on FranchiseID
        cursor.execute("""
            SELECT genericattribute6, genericattribute7 
            FROM ZRETAILER 
            WHERE genericattribute5 = ?
        """, (new_dim0,))
        result = cursor.fetchone()

        if result:
            new_city, new_region = result
        else:
            flash("Invalid FranchiseID.")
            return redirect(url_for('indexZmdlt', mdltcellseq=mdltcellseq))

        cursor.execute("""
            UPDATE ZMDLT
            SET effectivestartdate = ?, effectiveenddate = ?, mdltname = ?, value = ?, 
                unittypeforvalue = 'PKR', dim0 = ?, dim1 = ?, modificationdate = ?, city = ?, region = ?, eventtype = ?
            WHERE mdltcellseq = ?
        """, (new_effectivestartdate, new_effectiveenddate, new_mdltname, new_value,
              new_dim0, new_dim1, new_modificationdate, new_city, new_region, new_eventtype, mdltcellseq))
        conn.commit()

        flash('Entry successfully updated.')
        return redirect(url_for('indexZmdlt'))

    cursor.execute("""
    SELECT mdltcellseq, effectivestartdate, effectiveenddate, mdltname, value, 
           dim0, dim1, modificationdate, city, region, eventtype
    FROM ZMDLT 
    WHERE mdltcellseq = ?
    """, (mdltcellseq,))
    entry = cursor.fetchone()

    if entry:
        # Format the dates to YYYY-MM-DD format
        entry = list(entry)
        entry[1] = entry[1].strftime('%Y-%m-%d')
        entry[2] = entry[2].strftime('%Y-%m-%d')

    cursor.close()
    conn.close()

    if not entry:
        flash('Entry not found.')
        return redirect(url_for('indexZmdlt'))

    return render_template('editZmdlt.html', entry=entry, franchises=franchises, products=products)


@app.route('/viewLookup/<int:mdltcellseq>', methods=['GET'])
def viewLookup(mdltcellseq):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Fetch the product's data
    cursor.execute("""
        SELECT EFFECTIVESTARTDATE, EFFECTIVEENDDATE, MDLTNAME, DIM0, DIM1, VALUE, UNITTYPEFORVALUE, MODIFICATIONDATE, CITY, REGION, EVENTTYPE
        FROM ZMDLT 
        WHERE mdltcellseq = ?
    """, (mdltcellseq,))
    lookup_data = cursor.fetchone()
    cursor.close()
    conn.close()

    if lookup_data is None:
        flash('Retailer not found.', 'error')
        return redirect(url_for('indexZmdlt'))

    return render_template('viewLookup.html', lookup=lookup_data)


@app.route('/deleteZmdlt/<int:mdltcellseq>', methods=['POST'])
def deleteZmdlt(mdltcellseq):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM zmdlt WHERE mdltcellseq = ?", (mdltcellseq,))
    conn.commit()
    cursor.close()
    conn.close()
    flash('Entry successfully deleted.')
    return redirect(url_for('indexZmdlt'))


#DISPLAY MDLT

@app.route('/ZmdltDisplay')
def ZmdltDisplay():
    search_query = request.args.get('search')
    search_field = request.args.get('field')
    page = request.args.get('page', 1, type=int)
    per_page = 8

    conn = get_db_connection()
    cursor = conn.cursor()

    if search_query and search_field:
        # Validate that the field is one of the allowed search fields
        if search_field in ['dim0', 'dim1', 'value']:
            query = f"SELECT COUNT(*) FROM ZMDLT WHERE {search_field.upper()} LIKE ?"
            cursor.execute(query, (f"%{search_query}%",))
            total = cursor.fetchone()[0]

            query = f"""
            SELECT MDLTCELLSEQ, EFFECTIVESTARTDATE, EFFECTIVEENDDATE, MDLTNAME, VALUE, UNITTYPEFORVALUE, DIM0, DIM1, MODIFICATIONDATE
            FROM ZMDLT
            WHERE {search_field.upper()} LIKE ?
            ORDER BY MDLTCELLSEQ ASC
            LIMIT ? OFFSET ?
            """
            cursor.execute(query, (f"%{search_query}%", per_page, (page - 1) * per_page))
        else:
            flash('Invalid search field selected.')
            return redirect(url_for('indexZmdlt'))
    else:
        query = "SELECT COUNT(*) FROM ZMDLT"
        cursor.execute(query)
        total = cursor.fetchone()[0]

        query = """
        SELECT MDLTCELLSEQ, EFFECTIVESTARTDATE, EFFECTIVEENDDATE, MDLTNAME, VALUE, UNITTYPEFORVALUE, DIM0, DIM1, MODIFICATIONDATE
        FROM ZMDLT
        ORDER BY MDLTCELLSEQ ASC
        LIMIT ? OFFSET ?
        """
        cursor.execute(query, (per_page, (page - 1) * per_page))

    data = cursor.fetchall()
    cursor.close()
    conn.close()

    total_pages = math.ceil(total / per_page)
    
    return render_template('ZmdltDisplay.html', data=data, page=page, total_pages=total_pages, search_query=search_query)





# Define your routes here...

@app.route('/index_commission')
def index_commission():
    return render_template('index_commission.html')

@app.route('/calculate_commission', methods=['POST'])
def calculate_commission():
    p_dim1 = request.form['p_dim1']
    p_franchise_id = request.form['p_franchise_id']
    p_retailerid = request.form['p_retailerid']
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Hardcoded values
        p_retailer_msisdn = '3495255217'
        p_consumer_msisdn = '923222224557'
        p_date = '2024-07-18'
        p_comments = 'NILL'
        p_orderid = '93'
        p_eventtype = '1'
        p_linenumber = '3'
        p_sublinenumber = '2'

        # Prepare the call statement
        call_proc = f"""
        CALL MSHAFIQ.GET_NET_COMMISSION (
            '{p_retailerid}',
            '{p_dim1}',
            '{p_retailer_msisdn}',
            '{p_consumer_msisdn}',
            '{p_franchise_id}',
            '{p_date}',
            '{p_comments}',
            '{p_orderid}',
            '{p_eventtype}',
            '{p_linenumber}',
            '{p_sublinenumber}',
            ?)
        """

        # Execute the stored procedure
        cursor.execute(call_proc)
        
        # Fetch the results
        result_table = cursor.fetchall()

        # Assuming the result is a single row, single column with the net commission
        if result_table:
            net_commission = result_table[0][15]  # Adjust index based on your result_table structure
            flash(f'The calculated net commission is: {net_commission:.2f}', 'success')
        else:
            flash('No commission calculated. Please check your inputs.', 'error')
        
        cursor.close()
        conn.close()

    except Exception as e:
        flash(f'An error occurred: {str(e)}', 'error')

    return redirect(url_for('index_commission'))





# # Manual commissions
# def send_result_to_ers(salestransactionseq, RetailerMSISDN, netcommission):
#     try:
#         logging.info(f"ERS FUNC DETAILS: {salestransactionseq},{RetailerMSISDN},{netcommission}")
#         url = "http://10.96.2.36:9105/esb/rs/retailercreditdebit"
#         payload = json.dumps({
#             "requestHeader": {
#                 "requestId": salestransactionseq,
#                 "correlationId": "",
#                 "sourceSystem": "ERS",
#                 "timestamp": ""
#             },
#             "requestBody": {
#                 "msisdn": RetailerMSISDN,
#                 "stock0": str(netcommission),
#                 "reason": "credit",
#                 "commFlag": "0"
#             }
#         })
#         headers = {
#             'Content-Type': 'application/json'
#         }
#         username = 'admin'
#         password = 'admin'
#         auth = (username, password)
#         response = requests.post(url, headers=headers, data=payload, auth=auth)
#         logging.info(f"Received response from ERS for transaction {salestransactionseq}")
#         logging.info(response.json())
#         return response.json()
#     except Exception as e:
#         logging.info(f"An error occurred in ERS function: {e}")
#         return 'ERRORINERS'

# # Fetch data to send to ERS
# def fetch_data_from_db(cursor):
#     try:
#         query = "SELECT COMMID, MSISDN, NET_COMMISSION FROM ZMANUAL_COMMISSION WHERE STATUS = 'READY'"
#         cursor.execute(query)
#         result = cursor.fetchone()
#         if result:
#             return result  # Returns a tuple (COMMID, MSISDN, NET_COMMISSION)
#         else:
#             logging.info(f"No data found for STATUS: READY")
#             return None
#     except Exception as e:
#         logging.info(f"Database fetch error: {e}")
#         return None

# # Update commission status
# def update_commission_status(cursor, comm_id, correlation_id, error_msg, resultCode):
#     try:
#         params1 = (comm_id, correlation_id, error_msg, resultCode)
#         cursor.callproc('ers_manual_commission', params1)
#         cursor.connection.commit()
#     except Exception as e:
#         logging.info(f"Error updating commission status: {e}")

# # Background processing function
# def process_commissions(df):
#     # Connect to the database
#     conn = get_db_connection()
#     cursor = conn.cursor()

#     # Prepare batch insert query
#     insert_values = []
#     for index, row in df.iterrows():
#         # Add values to insert list
#         insert_values.append((
#             row['Business_Partner_Code'],  # Maps to RETAILERID
#             0,  # MSISDN - set to 0 for now
#             row['Commission_Type'],  # Maps to COMMISSION_TYPE
#             row['Adjustment Date'],  # Maps to ADJUSTMENT_DATE
#             row['Description'],  # Maps to DESCRIPTION
#             row['Value'],  # Maps to VALUE
#             row['Tax_Rate'],  # Maps to TAX_RATE
#             0,  # GST set to 0
#             0,  # GST_WITHHELD set to 0
#             0,  # WHT_TAX set to 0
#             0,  # NET_COMMISSION set to 0
#             'READY',  # STATUS set to 'READY'
#             None,  # CORELATION_ID set to None
#             None   # ERROR set to None
#         ))

#     # Execute batch insert
#     query = """
#         INSERT INTO ZMANUAL_COMMISSION (
#             RETAILERID, MSISDN, COMMISSION_TYPE, ADJUSTMENT_DATE, DESCRIPTION, VALUE, TAX_RATE, GST, GST_WITHHELD, WHT_TAX, NET_COMMISSION, STATUS, CORELATION_ID, ERROR
#         ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
#     """
#     cursor.executemany(query, insert_values)
#     conn.commit()

#     # Now, call the stored procedure for each commission
#     cursor.execute("SELECT COMMID, VALUE ,RETAILERID, TAX_RATE FROM ZMANUAL_COMMISSION WHERE STATUS = 'READY'")
#     result = cursor.fetchall()

#     for row in result:
#         commid, value, retailerid, tax_rate = row
#         params1 = (commid, retailerid, value, tax_rate)
#         cursor.callproc('get_net_manual_commission', params1)

#     conn.commit()

#     # Fetch data to send to ERS and process each row
#     while True:
#         data = fetch_data_from_db(cursor)
#         if not data:
#             break  # No more records to process
#         fetched_comm_id, fetched_msisdn, fetched_netcommission = data

#         # Start a thread to send the result to ERS
#         ersresult = send_result_to_ers(fetched_comm_id, fetched_msisdn, fetched_netcommission)
#         logging.info(f"ERS RESULT: {ersresult}")

#         # Initialize defaults
#         correlationId = None
#         resultCode = None
#         errorMsg = None

#         if ersresult == "ERRORINERS":
#             resultCode = "ERRORINERS"
#             errorMsg = "ERS Unreachable"
#         else:
#             resultCode = ersresult.get('resultHeader', {}).get('resultCode', 'UNKNOWN')
#             correlationId = ersresult.get('resultHeader', {}).get('correlationId', 'UNKNOWN')
#             errorMsg = ersresult.get('resultHeader', {}).get('externalErrorMessage', '')

#         logging.info(f'Correlation ID: {correlationId}, Result Code: {resultCode}')
#         if resultCode:
#             update_commission_status(cursor, fetched_comm_id, correlationId, errorMsg, resultCode)
#         else:
#             logging.info("Unable to proceed with database update due to missing resultCode.")

#     cursor.close()
#     conn.close()

# # Route for uploading Excel file
# @app.route('/upload_commissions', methods=['POST'])
# def upload_commissions():
#     if 'file' not in request.files:
#         flash('No file part')
#         return redirect(url_for('manual_commissions'))

#     file = request.files['file']
#     if file.filename == '':
#         flash('No selected file')
#         return redirect(url_for('manual_commissions'))

#     if file and file.filename.endswith('.xlsx'):
#         try:
#             # Read the Excel file into a pandas DataFrame
#             df = pd.read_excel(file)

#             # Expected headers
#             required_headers = ['Business_Partner_Code', 'Commission_Type', 'Adjustment Date', 'Description', 'Value', 'Tax_Rate']
#             if list(df.columns) != required_headers:
#                 flash('Incorrect headers in Excel file.')
#                 return redirect(url_for('manual_commissions'))

#             df['Tax_Rate'].fillna(0, inplace=True)

#             # Store the uploaded data in the session for immediate display
#             session['uploaded_data'] = df.to_dict(orient='records')

#             # Start the background processing in a new thread
#             thread = Thread(target=process_commissions, args=(df,))
#             thread.start()

#             flash('File uploaded successfully. Processing data in the background.')
#             return redirect(url_for('manual_commissions'))
#         except Exception as e:
#             flash(f'Error processing file: {e}')
#             return redirect(url_for('manual_commissions'))

#     flash('Invalid file format. Please upload an Excel file (.xlsx).')
#     return redirect(url_for('manual_commissions'))

# # Route for displaying the table
# @app.route('/manual_commissions', methods=['GET'])
# def manual_commissions():
#     search_query = request.args.get('search')
#     search_field = request.args.get('field', 'COMMID')  # Default field: COMMID
#     page = request.args.get('page', 1, type=int)
#     per_page = 8

#     conn = get_db_connection()
#     cursor = conn.cursor()

#     if search_query:
#         query = f"SELECT COUNT(*) FROM ZMANUAL_COMMISSION WHERE {search_field} LIKE ?"
#         cursor.execute(query, (f"%{search_query}%",))
#         total = cursor.fetchone()[0]

#         query = f"""
#         SELECT COMMID, RETAILERID, MSISDN, COMMISSION_TYPE, ADJUSTMENT_DATE, DESCRIPTION, VALUE, TAX_RATE, GST, GST_WITHHELD, WHT_TAX, NET_COMMISSION, STATUS, CORELATION_ID, ERROR
#         FROM ZMANUAL_COMMISSION
#         WHERE {search_field} LIKE ?
#         ORDER BY COMMID DESC
#         LIMIT ? OFFSET ?
#         """
#         cursor.execute(query, (f"%{search_query}%", per_page, (page - 1) * per_page))
#     else:
#         query = "SELECT COUNT(*) FROM ZMANUAL_COMMISSION"
#         cursor.execute(query)
#         total = cursor.fetchone()[0]

#         query = """
#         SELECT COMMID, RETAILERID, MSISDN, COMMISSION_TYPE, ADJUSTMENT_DATE, DESCRIPTION, VALUE, TAX_RATE, GST, GST_WITHHELD, WHT_TAX, NET_COMMISSION, STATUS, CORELATION_ID, ERROR
#         FROM ZMANUAL_COMMISSION
#         ORDER BY COMMID DESC
#         LIMIT ? OFFSET ?
#         """
#         cursor.execute(query, (per_page, (page - 1) * per_page))

#     data = cursor.fetchall()
#     cursor.close()
#     conn.close()

#     total_pages = math.ceil(total / per_page)

#     # Retrieve uploaded data from session for display
#     uploaded_data = session.pop('uploaded_data', None)

#     return render_template('manual_commissions.html', data=data, page=page, total_pages=total_pages, search_query=search_query, search_field=search_field, uploaded_data=uploaded_data)
# Manual commissions
def send_result_to_ers(salestransactionseq, RetailerMSISDN, netcommission):
    try:
        logging.info(f"ERS FUNC DETAILS: {salestransactionseq},{RetailerMSISDN},{netcommission}")
        url = "http://10.96.2.36:9105/esb/rs/retailercreditdebit"
        payload = json.dumps({
            "requestHeader": {
                "requestId": salestransactionseq,
                "correlationId": "",
                "sourceSystem": "ERS",
                "timestamp": ""
            },
            "requestBody": {
                "msisdn": RetailerMSISDN,
                "stock0": str(netcommission),
                "reason": "credit",
                "commFlag": "0"
            }
        })
        headers = {
            'Content-Type': 'application/json'
        }
        username = 'admin'
        password = 'admin'
        auth = (username, password)
        response = requests.post(url, headers=headers, data=payload, auth=auth)
        logging.info(f"Received response from ERS for transaction {salestransactionseq}")
        logging.info(response.json())
        return response.json()
    except Exception as e:
        logging.info(f"An error occurred in ERS function: {e}")
        return 'ERRORINERS'

# Fetch data to send to ERS
def fetch_data_from_db(cursor):
    try:
        query = "SELECT COMMID, MSISDN, NET_COMMISSION FROM ZMANUAL_COMMISSION WHERE STATUS = 'READY'"
        cursor.execute(query)
        result = cursor.fetchone()
        if result:
            return result  # Returns a tuple (COMMID, MSISDN, NET_COMMISSION)
        else:
            logging.info(f"No data found for STATUS: READY")
            return None
    except Exception as e:
        logging.info(f"Database fetch error: {e}")
        return None

# Update commission status
def update_commission_status(cursor, comm_id, correlation_id, error_msg, resultCode):
    try:
        params1 = (comm_id, correlation_id, error_msg, resultCode)
        cursor.callproc('ers_manual_commission', params1)
        cursor.connection.commit()
    except Exception as e:
        logging.info(f"Error updating commission status: {e}")

# Route for uploading Excel file
@app.route('/upload_commissions', methods=['POST'])
def upload_commissions():
    if 'file' not in request.files:
        flash('No file part')
        return redirect(url_for('manual_commissions'))

    file = request.files['file']
    if file.filename == '':
        flash('No selected file')
        return redirect(url_for('manual_commissions'))

    if file and file.filename.endswith('.xlsx'):
        try:
            # Read the Excel file into a pandas DataFrame
            df = pd.read_excel(file)

            # Expected headers
            required_headers = ['Business_Partner_Code', 'Commission_Type', 'Adjustment Date', 'Description', 'Value', 'Tax_Rate']
            if list(df.columns) != required_headers:
                flash('Incorrect headers in Excel file.')
                return redirect(url_for('manual_commissions'))

            df['Tax_Rate'].fillna(0, inplace=True)

            # Connect to the database
            conn = get_db_connection()
            cursor = conn.cursor()

            # Fetch all valid commission types and their GENERICATTRIBUTE1 from the ZCOMMISSION_TYPE table
            cursor.execute("SELECT COMMISSION_TYPE, GENERICATTRIBUTE1 FROM ZCOMMISSION_TYPE")
            commission_type_data = cursor.fetchall()

            # Create a dictionary with commission_type as key and GENERICATTRIBUTE1 as value (handling it as a string)
            commission_type_map = {row['COMMISSION_TYPE']: row['GENERICATTRIBUTE1'] for row in commission_type_data}

            # Prepare batch insert query
            insert_values = []
            for index, row in df.iterrows():
                commission_type = row['Commission_Type']

                # Validate if the Commission_Type exists in ZCOMMISSION_TYPE and GENERICATTRIBUTE1 is '1'
                if commission_type not in commission_type_map:
                    flash(f"The commission type in row {index + 2} does not exist in ZCOMMISSION_TYPE Table.")
                    return redirect(url_for('manual_commissions'))
                elif commission_type_map[commission_type] == '0':  # Compare as string '0'
                    flash(f"You have Uploaded unactive commission type in row {index + 2}.")
                    continue  # Skip this record

                # Add values to insert list
                insert_values.append((
                    row['Business_Partner_Code'],  # Maps to RETAILERID
                    0,  # MSISDN - set to 0 for now
                    commission_type,  # Maps to COMMISSION_TYPE
                    row['Adjustment Date'],  # Maps to ADJUSTMENT_DATE
                    row['Description'],  # Maps to DESCRIPTION
                    row['Value'],  # Maps to VALUE
                    row['Tax_Rate'],  # Maps to TAX_RATE
                    0,  # GST set to 0
                    0,  # GST_WITHHELD set to 0
                    0,  # WHT_TAX set to 0
                    0,  # NET_COMMISSION set to 0
                    'READY',  # STATUS set to 'READY'
                    None,  # CORELATION_ID set to None
                    None   # ERROR set to None
                ))

            # Execute batch insert
            query = """
                INSERT INTO ZMANUAL_COMMISSION (
                    RETAILERID, MSISDN, COMMISSION_TYPE, ADJUSTMENT_DATE, DESCRIPTION, VALUE, TAX_RATE, GST, GST_WITHHELD, WHT_TAX, NET_COMMISSION, STATUS, CORELATION_ID, ERROR
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            cursor.executemany(query, insert_values)
            conn.commit()
            

            # Now, call the stored procedure for each commission
            cursor.execute("SELECT COMMID, VALUE ,RETAILERID, TAX_RATE FROM ZMANUAL_COMMISSION WHERE STATUS = 'READY'")
            result = cursor.fetchall()

            for row in result:
                commid, value, retailerid, tax_rate = row
                params1 = (commid, retailerid, value, tax_rate)
                cursor.callproc('get_net_manual_commission', params1)

            conn.commit()

            # Fetch data to send to ERS and process each row
            while True:
                data = fetch_data_from_db(cursor)
                if not data:
                    break  # No more records to process
                fetched_comm_id, fetched_msisdn, fetched_netcommission = data

                # Start a thread to send the result to ERS
                ersresult = send_result_to_ers(fetched_comm_id, fetched_msisdn, fetched_netcommission)
                logging.info(f"ERS RESULT: {ersresult}")

                # Initialize defaults
                correlationId = None
                resultCode = None
                errorMsg = None

                if ersresult == "ERRORINERS":
                    resultCode = "ERRORINERS"
                    errorMsg = "ERS Unreachable"
                else:
                    resultCode = ersresult.get('resultHeader', {}).get('resultCode', 'UNKNOWN')
                    correlationId = ersresult.get('resultHeader', {}).get('correlationId', 'UNKNOWN')
                    errorMsg = ersresult.get('resultHeader', {}).get('externalErrorMessage', '')

                logging.info(f'Correlation ID: {correlationId}, Result Code: {resultCode}')
                if resultCode:
                    update_commission_status(cursor, fetched_comm_id, correlationId, errorMsg, resultCode)
                else:
                    logging.info("Unable to proceed with database update due to missing resultCode.")

            cursor.close()
            conn.close()

            flash('File uploaded and commissions processed successfully.')
            return redirect(url_for('manual_commissions'))

        except Exception as e:
            flash(f'Error processing file: {e}')
            return redirect(url_for('manual_commissions'))

    flash('Invalid file format. Please upload an Excel file (.xlsx).')
    return redirect(url_for('manual_commissions'))


# Route for displaying the table
@app.route('/manual_commissions', methods=['GET'])
def manual_commissions():
    search_query = request.args.get('search')
    search_field = request.args.get('field', 'COMMID')  # Default field: COMMID
    page = request.args.get('page', 1, type=int)
    per_page = 8

    conn = get_db_connection()
    cursor = conn.cursor()

    if search_query:
        query = f"SELECT COUNT(*) FROM ZMANUAL_COMMISSION WHERE {search_field} LIKE ?"
        cursor.execute(query, (f"%{search_query}%",))
        total = cursor.fetchone()[0]

        query = f"""
        SELECT COMMID, RETAILERID, MSISDN, COMMISSION_TYPE, ADJUSTMENT_DATE, DESCRIPTION, VALUE, TAX_RATE, GST, GST_WITHHELD, WHT_TAX, NET_COMMISSION, STATUS, CORELATION_ID, ERROR
        FROM ZMANUAL_COMMISSION
        WHERE {search_field} LIKE ?
        ORDER BY COMMID DESC
        LIMIT ? OFFSET ?
        """
        cursor.execute(query, (f"%{search_query}%", per_page, (page - 1) * per_page))
    else:
        query = "SELECT COUNT(*) FROM ZMANUAL_COMMISSION"
        cursor.execute(query)
        total = cursor.fetchone()[0]

        query = """
        SELECT COMMID, RETAILERID, MSISDN, COMMISSION_TYPE, ADJUSTMENT_DATE, DESCRIPTION, VALUE, TAX_RATE, GST, GST_WITHHELD, WHT_TAX, NET_COMMISSION, STATUS, CORELATION_ID, ERROR
        FROM ZMANUAL_COMMISSION
        ORDER BY COMMID DESC
        LIMIT ? OFFSET ?
        """
        cursor.execute(query, (per_page, (page - 1) * per_page))

    data = cursor.fetchall()
    cursor.close()
    conn.close()

    total_pages = math.ceil(total / per_page)

    return render_template('manual_commissions.html', data=data, page=page, total_pages=total_pages, search_query=search_query, search_field=search_field)







#######################################################################################################################################################################################


#######################################################################################################################################################################################


#######################################################################################################################################################################################




#######################################################################################################################################################################################




#NEW CAMPAIGN ROUTES

# TB Sales Transaction Error Table

@app.route('/TbTransactionsError')
def indexTbTransactionsError():
    search_query = request.args.get('search')
    search_field = request.args.get('field', 'salestransactionseq')  # Default search field
    page = request.args.get('page', 1, type=int)
    per_page = 8

    # Get download parameters
    download = request.args.get('download', False)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    conn = get_db_connection()
    cursor = conn.cursor()

    # List of valid search fields
    valid_fields = [
        'salestransactionseq', 'comments', 'compensationdate', 'error', 'eventtype', 'genericattribute1', 'genericattribute2',
        'genericattribute3', 'genericattribute6', 'genericattribute12', 'genericnumber1', 'genericnumber2', 'genericnumber3',
        'genericnumber4', 'genericnumber5', 'linenumber', 'orderid', 'origintypeid', 'productdescription', 'productid', 'productname',
        'sublinenumber', 'value', 'genericattribute27', 'genericattribute28', 'genericattribute29', 'genericattribute30', 'channel',
        'genericattribute22', 'genericattribute23', 'genericattribute24', 'genericattribute31', 'genericboolean1', 'genericboolean2'
    ]

    # Query for counting total records based on search criteria
    if search_query and search_field in valid_fields:
        count_query = f"SELECT COUNT(*) FROM z_tb_salestransaction_error WHERE {search_field} LIKE ?"
        params = (f"%{search_query}%",)
        cursor.execute(count_query, params)
        total = cursor.fetchone()[0]

        # Data query to retrieve paginated results based on search query
        data_query = f"""
            SELECT salestransactionseq, comments, compensationdate, error, eventtype, genericattribute1, genericattribute2,
                   genericattribute3, genericattribute6, genericattribute12, genericnumber1, genericnumber2, genericnumber3,
                   genericnumber4, genericnumber5, linenumber, orderid, origintypeid, productdescription, productid, productname,
                   sublinenumber, value, genericattribute27, genericattribute28, genericattribute29, genericattribute30, channel,
                   genericattribute22, genericattribute23, genericattribute24, genericattribute31, genericboolean1, genericboolean2
            FROM z_tb_salestransaction_error
            WHERE {search_field} LIKE ?
            ORDER BY CAST(salestransactionseq AS INTEGER) DESC
            LIMIT ? OFFSET ?
        """
        cursor.execute(data_query, (f"%{search_query}%", per_page, (page - 1) * per_page))
    else:
        # Query for counting total records without search query
        count_query = "SELECT COUNT(*) FROM z_tb_salestransaction_error"
        cursor.execute(count_query)
        total = cursor.fetchone()[0]

        # Data query to retrieve paginated results
        data_query = """
            SELECT salestransactionseq, comments, compensationdate, error, eventtype, genericattribute1, genericattribute2,
                   genericattribute3, genericattribute6, genericattribute12, genericnumber1, genericnumber2, genericnumber3,
                   genericnumber4, genericnumber5, linenumber, orderid, origintypeid, productdescription, productid, productname,
                   sublinenumber, value, genericattribute27, genericattribute28, genericattribute29, genericattribute30, channel,
                   genericattribute22, genericattribute23, genericattribute24, genericattribute31, genericboolean1, genericboolean2
            FROM z_tb_salestransaction_error
            ORDER BY CAST(salestransactionseq AS INTEGER) DESC
            LIMIT ? OFFSET ?
        """
        cursor.execute(data_query, (per_page, (page - 1) * per_page))

    data = cursor.fetchall()
    cursor.close()
    conn.close()

    total_pages = math.ceil(total / per_page)

    # If download is requested, fetch and export all data to Excel
    if download:
        try:
            return export_tb_sales_transactions_error_to_excel(search_query, search_field, start_date, end_date)
        except ValueError as ve:
            flash(str(ve), 'danger')
            return redirect(url_for('indexTbTransactionsError', search=search_query, field=search_field))
        except Exception as e:
            app.logger.error(f"Error exporting data: {e}")
            flash("An unexpected error occurred. Please try again later.", 'danger')
            return redirect(url_for('indexTbTransactionsError', search=search_query, field=search_field))

    return render_template(
        'indexTbTransactionsError.html',
        data=data,
        page=page,
        total_pages=total_pages,
        search_query=search_query,
        search_field=search_field
    )


# Function to fetch all data from the database for exporting to Excel
def fetch_tb_sales_transactions_error_data(search_query=None, search_field=None, start_date=None, end_date=None):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Base query to select data from z_tb_salestransaction_error
    query = """
        SELECT salestransactionseq, comments, compensationdate, error, eventtype, genericattribute1, genericattribute2,
               genericattribute3, genericattribute6, genericattribute12, genericnumber1, genericnumber2, genericnumber3,
               genericnumber4, genericnumber5, linenumber, orderid, origintypeid, productdescription, productid, productname,
               sublinenumber, value, genericattribute27, genericattribute28, genericattribute29, genericattribute30, channel,
               genericattribute22, genericattribute23, genericattribute24, genericattribute31, genericboolean1, genericboolean2
        FROM z_tb_salestransaction_error
    """

    params = []
    conditions = []

    # Add search query if provided
    if search_query and search_field:
        conditions.append(f"{search_field} LIKE ?")
        params.append(f"%{search_query}%")

    # Add date range filter if start_date and end_date are provided
    if start_date and end_date:
        conditions.append("compensationdate BETWEEN ? AND ?")
        params.extend([start_date, end_date])

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    cursor.execute(query, params)
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return data


# Function to stream Excel data as it's generated for download
def generate_tb_excel_stream_for_error_data(data, headers, max_rows_per_sheet=1000000):
    # Create an in-memory Excel file
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})

    total_rows = len(data)
    sheet_number = 1

    for start_idx in range(0, total_rows, max_rows_per_sheet):
        end_idx = min(start_idx + max_rows_per_sheet, total_rows)
        sheet_data = data[start_idx:end_idx]

        # Add new sheet
        worksheet = workbook.add_worksheet(f"Sheet{sheet_number}")
        sheet_number += 1

        # Write headers to the first row of the sheet
        worksheet.write_row(0, 0, headers)

        # Write data row by row starting from the second row (index 1)
        for row_idx, row_data in enumerate(sheet_data, start=1):
            worksheet.write_row(row_idx, 0, row_data)

    workbook.close()
    output.seek(0)

    return output.getvalue()


@app.route('/download_tb_sales_transactions_error_data', methods=['GET'])
def export_tb_sales_transactions_error_to_excel(search_query=None, search_field=None, start_date=None, end_date=None):
    # Fetch data from the database
    data = fetch_tb_sales_transactions_error_data(search_query, search_field, start_date, end_date)
    
    headers = [
        "salestransactionseq", "comments", "compensationdate", "error", "eventtype", "genericattribute1", "genericattribute2",
        "genericattribute3", "genericattribute6", "genericattribute12", "genericnumber1", "genericnumber2", "genericnumber3",
        "genericnumber4", "genericnumber5", "linenumber", "orderid", "origintypeid", "productdescription", "productid", "productname",
        "sublinenumber", "value", "genericattribute27", "genericattribute28", "genericattribute29", "genericattribute30", "channel",
        "genericattribute22", "genericattribute23", "genericattribute24", "genericattribute31", "genericboolean1", "genericboolean2"
    ]

    # Stream Excel file as it's generated
    def generate():
        yield generate_tb_excel_stream_for_error_data(data, headers)

    return Response(generate(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=tberror_transactions.xlsx"})



@app.route('/TbviewErrors/<int:salestransactionseq>', methods=['GET'])
def TbviewErrors(salestransactionseq):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Fetch the product's data
    cursor.execute("""
        SELECT salestransactionseq, comments, compensationdate, error, eventtype, genericattribute1, genericattribute2, genericattribute3, genericattribute6, genericattribute12, genericnumber1, genericnumber2, genericnumber3, genericnumber4, genericnumber5, linenumber, orderid, origintypeid, productdescription, productid, productname, sublinenumber, value, genericattribute27, genericattribute28, genericattribute29, genericattribute30, channel, genericattribute22, genericattribute23, genericattribute24, genericattribute31, genericboolean1, genericboolean2
        FROM z_tb_salestransaction_error 
        WHERE salestransactionseq = ?
    """, (salestransactionseq,))
    error_data = cursor.fetchone()
    cursor.close()
    conn.close()

    if error_data is None:
        flash('Retailer not found.', 'error')
        return redirect(url_for('indexTbTransactionsError'))

    return render_template('TbviewErrors.html', error=error_data)




@app.route('/TbeditTransactions/<int:salestransactionseq>', methods=['GET', 'POST'])
def TbeditTransactionsError(salestransactionseq):
    conn = get_db_connection()
    cursor = conn.cursor()
    if request.method == 'POST':
        # Fetch form data and handle potential missing data
        try:
            new_comments = request.form['comments']
            new_compensationdate = request.form['compensationdate']
            new_error = request.form['error']
            new_eventtype = request.form['eventtype']
            new_genericattribute1 = request.form['genericattribute1']
            new_genericattribute2 = request.form['genericattribute2']
            new_genericattribute3 = request.form['genericattribute3']
            new_genericattribute6 = request.form['genericattribute6']
            new_genericattribute12 = request.form['genericattribute12']
            new_genericnumber1 = request.form['genericnumber1']
            new_genericnumber2 = request.form['genericnumber2']
            new_genericnumber3 = request.form['genericnumber3']
            new_genericnumber4 = request.form['genericnumber4']
            new_genericnumber5 = request.form['genericnumber5']
            new_linenumber = request.form['linenumber']
            new_orderid = request.form['orderid']
            new_origintypeid = request.form['origintypeid']
            new_productdescription = request.form['productdescription']
            new_productid = request.form['productid']
            new_productname = request.form['productname']
            new_sublinenumber = request.form['sublinenumber']
            new_value = request.form['value']
            new_genericattribute27 = request.form['genericattribute27']
            new_genericattribute28 = request.form['genericattribute28']
            new_genericattribute29 = request.form['genericattribute29']
            new_genericattribute30 = request.form['genericattribute30']
       

            # Convert to integers where necessary
            new_linenumber = int(new_linenumber)
            new_sublinenumber = int(new_sublinenumber)
        except ValueError as ve:
            flash(f'Error with data: {ve}')
            return redirect(url_for('TbeditTransactionsError', salestransactionseq=salestransactionseq))
        except KeyError as ke:
            flash(f'Missing form field: {ke}')
            return redirect(url_for('TbeditTransactionsError', salestransactionseq=salestransactionseq))

        # Update the record
        try:
            cursor.execute("""
            UPDATE z_tb_salestransaction_error
            SET comments = ?, compensationdate = ?, error = ?, eventtype = ?, 
                genericattribute1 = ?, genericattribute2 = ?, genericattribute3 = ?, genericattribute6 = ?, genericattribute12 = ?,
                genericnumber1 = ?, genericnumber2 = ?, genericnumber3 = ?, genericnumber4 = ?, genericnumber5 = ?, linenumber = ?, orderid = ?, origintypeid = ?, productdescription = ?, 
                productid = ?, productname = ?, sublinenumber = ?, value = ?, 
                genericattribute27 = ?, genericattribute28 = ?, genericattribute29 = ?, genericattribute30 = ?
            WHERE salestransactionseq = ?
            """, (new_comments, new_compensationdate, new_error, new_eventtype, new_genericattribute1, 
                  new_genericattribute2, new_genericattribute3, new_genericattribute6, new_genericattribute12, new_genericnumber1, 
                  new_genericnumber2, new_genericnumber3, new_genericnumber4, new_genericnumber5, 
                  new_linenumber, new_orderid, new_origintypeid, new_productdescription, 
                  new_productid, new_productname, new_sublinenumber, new_value, new_genericattribute27, new_genericattribute28, new_genericattribute29, new_genericattribute30, salestransactionseq))
            conn.commit()
            flash('Entry successfully updated.')
        except dbapi.Error as e:
            flash(f'Database error: {e}')
        finally:
            cursor.close()
            conn.close()
 
        return redirect(url_for('indexTbTransactionsError'))
    
    cursor.execute("""
    SELECT  salestransactionseq, comments, compensationdate, error, eventtype, genericattribute1, genericattribute2, genericattribute3, genericattribute6, genericattribute12, genericnumber1, genericnumber2, genericnumber3, genericnumber4, genericnumber5, linenumber, orderid, origintypeid, productdescription, productid, productname, sublinenumber, value, genericattribute27, genericattribute28, genericattribute29, genericattribute30
    FROM z_tb_salestransaction_error 
    WHERE salestransactionseq = ?
    """, (salestransactionseq,))
    entry = cursor.fetchone()
    cursor.close()
    conn.close()

    if not entry:
        flash('Entry not found.')
        return redirect(url_for('indexTbTransactionsError'))

    return render_template('TbeditTransactionsError.html', entry=entry)


@app.route('/deleteTbTransactions/<int:salestransactionseq>', methods=['POST'])
def deleteTbTransactionsError(salestransactionseq):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM z_tb_salestransaction_error WHERE salestransactionseq = ?", (salestransactionseq,))
    conn.commit()
    cursor.close()
    conn.close()
    flash('Entry successfully deleted.')
    return redirect(url_for('indexTbTransactionsError'))





# TB Sales Transaction ERS Table

@app.route('/TbSalesTransaction')
def indexTbSalesTransaction():
    search_query = request.args.get('search')
    search_field = request.args.get('field', 'salestransactionseq')  # Default search field
    page = request.args.get('page', 1, type=int)
    per_page = 8

    # Get download parameters
    download = request.args.get('download', False)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    conn = get_db_connection()
    cursor = conn.cursor()

    # List of valid search fields
    valid_fields = [
        'salestransactionseq', 'comments', 'compensationdate', 'error_msg', 'eventtype', 'genericattribute1', 'genericattribute2',
        'genericattribute3', 'genericattribute6', 'genericattribute12', 'genericnumber1', 'genericnumber2', 'genericnumber3',
        'genericnumber4', 'genericnumber5', 'linenumber', 'orderid', 'origintypeid', 'productdescription', 'productid', 'productname',
        'sublinenumber', 'value', 'genericattribute27', 'genericattribute28', 'genericattribute29', 'genericattribute30', 'channel',
        'genericattribute22', 'genericattribute23', 'genericattribute24', 'genericattribute31', 'genericboolean1', 'genericboolean2'
    ]

    # Query for counting total records based on search criteria
    if search_query and search_field in valid_fields:
        count_query = f"SELECT COUNT(*) FROM Z_TB_SALESTRANSACTION_ERS WHERE {search_field} LIKE ?"
        params = (f"%{search_query}%",)
        cursor.execute(count_query, params)
        total = cursor.fetchone()[0]

        # Data query to retrieve paginated results based on search query
        data_query = f"""
            SELECT salestransactionseq, comments, compensationdate, error_msg, eventtype, genericattribute1, genericattribute2,
                   genericattribute3, genericattribute6, genericattribute12, genericnumber1, genericnumber2, genericnumber3,
                   genericnumber4, genericnumber5, linenumber, orderid, origintypeid, productdescription, productid, productname,
                   sublinenumber, value, genericattribute27, genericattribute28, genericattribute29, genericattribute30, channel,
                   genericattribute22, genericattribute23, genericattribute24, genericattribute31, genericboolean1, genericboolean2
            FROM Z_TB_SALESTRANSACTION_ERS
            WHERE {search_field} LIKE ?
            ORDER BY CAST(salestransactionseq AS INTEGER) DESC
            LIMIT ? OFFSET ?
        """
        cursor.execute(data_query, (f"%{search_query}%", per_page, (page - 1) * per_page))
    else:
        # Query for counting total records without search query
        count_query = "SELECT COUNT(*) FROM Z_TB_SALESTRANSACTION_ERS"
        cursor.execute(count_query)
        total = cursor.fetchone()[0]

        # Data query to retrieve paginated results
        data_query = """
            SELECT salestransactionseq, comments, compensationdate, error_msg, eventtype, genericattribute1, genericattribute2,
                   genericattribute3, genericattribute6, genericattribute12, genericnumber1, genericnumber2, genericnumber3,
                   genericnumber4, genericnumber5, linenumber, orderid, origintypeid, productdescription, productid, productname,
                   sublinenumber, value, genericattribute27, genericattribute28, genericattribute29, genericattribute30, channel,
                   genericattribute22, genericattribute23, genericattribute24, genericattribute31, genericboolean1, genericboolean2
            FROM Z_TB_SALESTRANSACTION_ERS
            ORDER BY CAST(salestransactionseq AS INTEGER) DESC
            LIMIT ? OFFSET ?
        """
        cursor.execute(data_query, (per_page, (page - 1) * per_page))

    data = cursor.fetchall()
    cursor.close()
    conn.close()

    total_pages = math.ceil(total / per_page)

    # If download is requested, fetch and export all data to Excel
    if download:
        try:
            return export_tb_sales_transactions_ers_to_excel(search_query, search_field, start_date, end_date)
        except ValueError as ve:
            flash(str(ve), 'danger')
            return redirect(url_for('indexTbSalesTransaction', search=search_query, field=search_field))
        except Exception as e:
            app.logger.error(f"Error exporting data: {e}")
            flash("An unexpected error occurred. Please try again later.", 'danger')
            return redirect(url_for('indexTbSalesTransaction', search=search_query, field=search_field))

    return render_template(
        'indexTbSalesTransaction.html',
        data=data,
        page=page,
        total_pages=total_pages,
        search_query=search_query,
        search_field=search_field
    )


# Function to fetch all data from the database for exporting to Excel
def fetch_tb_sales_transactions_ers_data(search_query=None, search_field=None, start_date=None, end_date=None):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Base query to select data from Z_TB_SALESTRANSACTION_ERS
    query = """
        SELECT salestransactionseq, comments, compensationdate, error_msg, eventtype, genericattribute1, genericattribute2,
               genericattribute3, genericattribute6, genericattribute12, genericnumber1, genericnumber2, genericnumber3,
               genericnumber4, genericnumber5, linenumber, orderid, origintypeid, productdescription, productid, productname,
               sublinenumber, value, genericattribute27, genericattribute28, genericattribute29, genericattribute30, channel,
               genericattribute22, genericattribute23, genericattribute24, genericattribute31, genericboolean1, genericboolean2
        FROM Z_TB_SALESTRANSACTION_ERS
    """

    params = []
    conditions = []

    # Add search query if provided
    if search_query and search_field:
        conditions.append(f"{search_field} LIKE ?")
        params.append(f"%{search_query}%")

    # Add date range filter if start_date and end_date are provided
    if start_date and end_date:
        conditions.append("compensationdate BETWEEN ? AND ?")
        params.extend([start_date, end_date])

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    cursor.execute(query, params)
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return data


# Function to stream Excel data as it's generated for download
def generate_tb_excel_stream_for_ers_data(data, headers, max_rows_per_sheet=1000000):
    # Create an in-memory Excel file
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})

    total_rows = len(data)
    sheet_number = 1

    for start_idx in range(0, total_rows, max_rows_per_sheet):
        end_idx = min(start_idx + max_rows_per_sheet, total_rows)
        sheet_data = data[start_idx:end_idx]

        # Add new sheet
        worksheet = workbook.add_worksheet(f"Sheet{sheet_number}")
        sheet_number += 1

        # Write headers to the first row of the sheet
        worksheet.write_row(0, 0, headers)

        # Write data row by row starting from the second row (index 1)
        for row_idx, row_data in enumerate(sheet_data, start=1):
            worksheet.write_row(row_idx, 0, row_data)

    workbook.close()
    output.seek(0)

    return output.getvalue()


@app.route('/download_tb_sales_transactions_ers_data', methods=['GET'])
def export_tb_sales_transactions_ers_to_excel(search_query=None, search_field=None, start_date=None, end_date=None):
    # Fetch data from the database
    data = fetch_tb_sales_transactions_ers_data(search_query, search_field, start_date, end_date)
    
    headers = [
        "salestransactionseq", "comments", "compensationdate", "error_msg", "eventtype", "genericattribute1", "genericattribute2",
        "genericattribute3", "genericattribute6", "genericattribute12", "genericnumber1", "genericnumber2", "genericnumber3",
        "genericnumber4", "genericnumber5", "linenumber", "orderid", "origintypeid", "productdescription", "productid", "productname",
        "sublinenumber", "value", "genericattribute27", "genericattribute28", "genericattribute29", "genericattribute30", "channel",
        "genericattribute22", "genericattribute23", "genericattribute24", "genericattribute31", "genericboolean1", "genericboolean2"
    ]

    # Stream Excel file as it's generated
    def generate():
        yield generate_tb_excel_stream_for_ers_data(data, headers)

    return Response(generate(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=tb_ers_transactions.xlsx"})





# TB_CAMPAIGN TABLE

# TB_CAMPAIGN TABLE

@app.route('/TbCampaign', methods=['GET', 'POST'])
def indexTbCampaign():
    if request.method == 'POST':
        excel_file = request.files.get('file')
        if excel_file and excel_file.filename.endswith('.xlsx'):
            try:
                # Read the Excel file
                df = pd.read_excel(excel_file)

                # Convert column names to upper case and strip spaces
                df.columns = df.columns.str.strip().str.upper()

                required_columns = [
                    'CAMPAIGNNAME', 'STARTDATE', 'ENDDATE', 'STATUS', 'TARGETCOUNT'
                ]

                missing_columns = [col for col in required_columns if col not in df.columns]
                extra_columns = [col for col in df.columns if col not in required_columns]

                if missing_columns or extra_columns:
                    error_message = "Error: The following issues were found in the Excel file:\n"
                    if missing_columns:
                        error_message += f"Missing columns: {', '.join(missing_columns)}\n"
                    if extra_columns:
                        error_message += f"Extra columns: {', '.join(extra_columns)}\n"
                    flash(error_message.strip())
                    return redirect(url_for('indexTbCampaign'))

                # Check for missing values in specific columns
                if df[['STARTDATE', 'ENDDATE', 'TARGETCOUNT']].isnull().any().any():
                    missing_values = df[['STARTDATE', 'ENDDATE', 'TARGETCOUNT']].isnull().sum()
                    error_message = "Error: The following fields have missing values:\n"
                    for field, count in missing_values.items():
                        if count > 0:
                            error_message += f"{field}: {count} missing values\n"
                    flash(error_message.strip())
                    return redirect(url_for('indexTbCampaign'))

                # Validate TARGETCOUNT as integer
                if not df['TARGETCOUNT'].apply(lambda x: isinstance(x, int)).all():
                    flash("Error: All TARGETCOUNT values must be integers.")
                    return redirect(url_for('indexTbCampaign'))

                # Validate STATUS as 0 or 1
                if not df['STATUS'].isin([0, 1]).all():
                    flash("Error: STATUS values must be either 0 or 1.")
                    return redirect(url_for('indexTbCampaign'))

                # Explicit date parsing
                df['STARTDATE'] = pd.to_datetime(df['STARTDATE'], errors='coerce')
                df['ENDDATE'] = pd.to_datetime(df['ENDDATE'], errors='coerce')

                # Remove rows where dates are invalid (NaT values)
                df.dropna(subset=['STARTDATE', 'ENDDATE'], inplace=True)

                # Check if STARTDATE is before ENDDATE
                for index, row in df.iterrows():
                    if row['STARTDATE'] >= row['ENDDATE']:
                        flash(f"Error: The STARTDATE in row {index + 2} must be before the ENDDATE.")
                        return redirect(url_for('indexTbCampaign'))

                conn = get_db_connection()
                cursor = conn.cursor()

                # Prepare data for insertion
                new_entries = []
                for index, row in df.iterrows():
                    # Check for duplicates
                    cursor.execute("""
                        SELECT COUNT(*) FROM Z_TB_CAMPAIGN
                        WHERE CAMPAIGNNAME = ? AND STARTDATE = ? AND ENDDATE = ? AND STATUS = ? AND TARGETCOUNT = ?
                    """, (row['CAMPAIGNNAME'], row['STARTDATE'], row['ENDDATE'], row['STATUS'], row['TARGETCOUNT']))
                    exists = cursor.fetchone()[0]

                    if exists == 0:
                        new_entries.append((
                            row['CAMPAIGNNAME'], row['STARTDATE'], row['ENDDATE'],
                            row['STATUS'], row['TARGETCOUNT']
                        ))

                # Batch insert new records
                if new_entries:
                    cursor.executemany("""
                        INSERT INTO Z_TB_CAMPAIGN (CAMPAIGNNAME, STARTDATE, ENDDATE, STATUS, TARGETCOUNT)
                        VALUES (?, ?, ?, ?, ?)
                    """, new_entries)

                conn.commit()
                cursor.close()
                conn.close()

                if new_entries:
                    flash('Excel file data successfully processed and added to the database.')
                else:
                    flash('No new entries were added as all records already exist in the database.')
            except Exception as e:
                flash(f'Error processing the file: {str(e)}')
        else:
            flash('Please upload a valid .xlsx file.')

        return redirect(url_for('indexTbCampaign'))

    search_query = request.args.get('search')
    search_field = request.args.get('field', 'CAMPAIGNNAME').upper()
    page = request.args.get('page', 1, type=int)
    per_page = 8

    conn = get_db_connection()
    cursor = conn.cursor()

    if search_query:
        query = f"SELECT COUNT(*) FROM Z_TB_CAMPAIGN WHERE {search_field} LIKE ?"
        cursor.execute(query, (f"%{search_query}%",))
        total = cursor.fetchone()[0]

        query = f"""
        SELECT CAMPAIGNID, CAMPAIGNNAME, STARTDATE, ENDDATE, STATUS, TARGETCOUNT
        FROM Z_TB_CAMPAIGN
        WHERE {search_field} LIKE ?
        ORDER BY CAMPAIGNID DESC
        LIMIT ? OFFSET ?
        """
        cursor.execute(query, (f"%{search_query}%", per_page, (page - 1) * per_page))
    else:
        query = "SELECT COUNT(*) FROM Z_TB_CAMPAIGN"
        cursor.execute(query)
        total = cursor.fetchone()[0]

        query = """
        SELECT CAMPAIGNID, CAMPAIGNNAME, STARTDATE, ENDDATE, STATUS, TARGETCOUNT
        FROM Z_TB_CAMPAIGN
        ORDER BY CAMPAIGNID DESC
        LIMIT ? OFFSET ?
        """
        cursor.execute(query, (per_page, (page - 1) * per_page))

    data = cursor.fetchall()
    cursor.close()
    conn.close()

    total_pages = math.ceil(total / per_page)

    if 'download' in request.args:
        return export_tb_campaign_to_excel(search_query, search_field)

    return render_template('indexTbCampaign.html', data=data, page=page, total_pages=total_pages, search_query=search_query, search_field=search_field)


def export_tb_campaign_to_excel(search_query=None, search_field=None):
    conn = get_db_connection()
    cursor = conn.cursor()

    batch_size = 10000
    offset = 0
    all_rows = []

    while True:
        if search_query:
            query = f"""
            SELECT CAMPAIGNNAME, STARTDATE, ENDDATE, STATUS, TARGETCOUNT
            FROM Z_TB_CAMPAIGN
            WHERE {search_field} LIKE ?
            ORDER BY CAMPAIGNID DESC
            LIMIT ? OFFSET ?
            """
            params = (f"%{search_query}%", batch_size, offset)
        else:
            query = """
            SELECT CAMPAIGNNAME, STARTDATE, ENDDATE, STATUS, TARGETCOUNT
            FROM Z_TB_CAMPAIGN
            ORDER BY CAMPAIGNID DESC
            LIMIT ? OFFSET ?
            """
            params = (batch_size, offset)

        cursor.execute(query, params)
        rows = cursor.fetchall()

        if not rows:
            break

        all_rows.extend(rows)
        offset += batch_size

    conn.close()

    df = pd.DataFrame(all_rows, columns=[
        'CampaignName', 'StartDate', 'EndDate', 'Status', 'TargetCount'
    ])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='TbCampaign')

    output.seek(0)

    return send_file(output, as_attachment=True, download_name='TbCampaignData.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/addTbCampaignEntry', methods=['GET', 'POST'])
def addTbCampaignEntry():
    if request.method == 'POST':
        campaignname = request.form['campaignname']
        startdate = request.form['startdate']
        enddate = request.form['enddate']
        status = request.form['status']
        targetcount = request.form['targetcount']

        # Validate date format
        try:
            startdate = datetime.strptime(startdate, '%Y-%m-%d')
            enddate = datetime.strptime(enddate, '%Y-%m-%d')
        except ValueError:
            flash("Invalid date format. Please use YYYY-MM-DD.")
            return redirect(url_for('indexTbCampaign'))

        # Check for missing values
        if not campaignname or not startdate or not enddate or not status or not targetcount:
            flash("All fields are required.")
            return redirect(url_for('indexTbCampaign'))

        # Validate TARGETCOUNT as integer
        if not targetcount.isdigit():
            flash("TARGETCOUNT must be an integer.")
            return redirect(url_for('indexTbCampaign'))

        # Validate STATUS as 0 or 1
        if status not in ['0', '1']:
            flash("STATUS must be either 0 or 1.")
            return redirect(url_for('indexTbCampaign'))

       

        conn = get_db_connection()
        cursor = conn.cursor()

        try:
            # Insert new entry
            cursor.execute("""
                INSERT INTO Z_TB_CAMPAIGN (CAMPAIGNNAME, STARTDATE, ENDDATE, STATUS, TARGETCOUNT)
                VALUES (?, ?, ?, ?, ?)
            """, (campaignname, startdate, enddate, status, targetcount))

            conn.commit()
            flash('New TbCampaign entry added successfully.')
        except Exception as e:
            flash(f'An error occurred: {str(e)}')
        finally:
            cursor.close()
            conn.close()

        return redirect(url_for('indexTbCampaign'))

    return render_template('addTbCampaignEntry.html')


@app.route('/editTbCampaign/<int:campaignid>', methods=['GET', 'POST'])
def editTbCampaign(campaignid):
    conn = get_db_connection()
    cursor = conn.cursor()

    if request.method == 'POST':
        new_campaignname = request.form['campaignname']
        new_startdate = request.form['startdate']
        new_enddate = request.form['enddate']
        new_status = request.form['status']
        new_targetcount = request.form['targetcount']

        # Validate date format
        try:
            new_startdate = datetime.strptime(new_startdate, '%Y-%m-%d')
            new_enddate = datetime.strptime(new_enddate, '%Y-%m-%d')
        except ValueError:
            flash("Invalid date format. Please use YYYY-MM-DD.")
            return redirect(url_for('editTbCampaign', campaignid=campaignid))

        cursor.execute("""
            UPDATE Z_TB_CAMPAIGN
            SET CAMPAIGNNAME = ?, STARTDATE = ?, ENDDATE = ?, STATUS = ?, TARGETCOUNT = ?
            WHERE CAMPAIGNID = ?
        """, (new_campaignname, new_startdate, new_enddate, new_status, new_targetcount, campaignid))
        conn.commit()

        flash('Entry successfully updated.')
        return redirect(url_for('indexTbCampaign'))

    cursor.execute("""
    SELECT CAMPAIGNID, CAMPAIGNNAME, STARTDATE, ENDDATE, STATUS, TARGETCOUNT
    FROM Z_TB_CAMPAIGN
    WHERE CAMPAIGNID = ?
    """, (campaignid,))
    entry = cursor.fetchone()

    if entry:
        # Format the dates to YYYY-MM-DD format
        entry = list(entry)
        entry[2] = entry[2].strftime('%Y-%m-%d')
        entry[3] = entry[3].strftime('%Y-%m-%d')

    cursor.close()
    conn.close()

    if not entry:
        flash('Entry not found.')
        return redirect(url_for('indexTbCampaign'))

    return render_template('editTbCampaign.html', entry=entry)


@app.route('/deleteTbCampaign/<int:campaignid>', methods=['POST'])
def deleteTbCampaign(campaignid):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM Z_TB_CAMPAIGN WHERE CAMPAIGNID = ?", (campaignid,))
    conn.commit()
    cursor.close()
    conn.close()
    flash('Entry successfully deleted.')
    return redirect(url_for('indexTbCampaign'))


@app.route('/viewTbCampaign/<int:campaignid>', methods=['GET'])
def viewTbCampaign(campaignid):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Fetch the campaign's data
    cursor.execute("""
        SELECT CAMPAIGNID, CAMPAIGNNAME, STARTDATE, ENDDATE, STATUS, TARGETCOUNT
        FROM Z_TB_CAMPAIGN
        WHERE CAMPAIGNID = ?
    """, (campaignid,))
    campaign_data = cursor.fetchone()
    cursor.close()
    conn.close()

    if campaign_data is None:
        flash('Campaign not found.', 'error')
        return redirect(url_for('indexTbCampaign'))

    # Format the dates to YYYY-MM-DD format for display
    campaign_data = list(campaign_data)
    campaign_data[2] = campaign_data[2].strftime('%Y-%m-%d')
    campaign_data[3] = campaign_data[3].strftime('%Y-%m-%d')

    return render_template('viewTbCampaign.html', campaign=campaign_data)


# Z_TB_COMMISSION_LOOKUP TABLE
@app.route('/TbCommissionLookup', methods=['GET', 'POST'])
def indexTbCommissionLookup():
    if request.method == 'POST':
        excel_file = request.files.get('file')
        if excel_file and excel_file.filename.endswith('.xlsx'):
            try:
                # Read the Excel file
                df = pd.read_excel(excel_file)

                # Convert column names to upper case and strip spaces
                df.columns = df.columns.str.strip().str.upper()

                required_columns = [
                    'RETAILERID', 'PRODUCTID', 'STARTDATE', 'ENDDATE', 'COMMISSIONVALUE'
                ]

                missing_columns = [col for col in required_columns if col not in df.columns]
                extra_columns = [col for col in df.columns if col not in required_columns]

                if missing_columns or extra_columns:
                    error_message = "Error: The following issues were found in the Excel file:\n"
                    if missing_columns:
                        error_message += f"Missing columns: {', '.join(missing_columns)}\n"
                    if extra_columns:
                        error_message += f"Extra columns: {', '.join(extra_columns)}\n"
                    flash(error_message.strip())
                    return redirect(url_for('indexTbCommissionLookup'))

                # Check for missing values in specific columns
                if df[['STARTDATE', 'ENDDATE', 'COMMISSIONVALUE']].isnull().any().any():
                    missing_values = df[['STARTDATE', 'ENDDATE', 'COMMISSIONVALUE']].isnull().sum()
                    error_message = "Error: The following fields have missing values:\n"
                    for field, count in missing_values.items():
                        if count > 0:
                            error_message += f"{field}: {count} missing values\n"
                    flash(error_message.strip())
                    return redirect(url_for('indexTbCommissionLookup'))

                # Validate RETAILERID and PRODUCTID as strings
                if not df['RETAILERID'].apply(lambda x: isinstance(x, str)).all():
                    flash("Error: All RETAILERID values must be strings.")
                    return redirect(url_for('indexTbCommissionLookup'))

                if not df['PRODUCTID'].apply(lambda x: isinstance(x, str)).all():
                    flash("Error: All PRODUCTID values must be strings.")
                    return redirect(url_for('indexTbCommissionLookup'))

                # Explicit date parsing
                df['STARTDATE'] = pd.to_datetime(df['STARTDATE'], errors='coerce')
                df['ENDDATE'] = pd.to_datetime(df['ENDDATE'], errors='coerce')

                # Remove rows where dates are invalid (NaT values)
                df.dropna(subset=['STARTDATE', 'ENDDATE'], inplace=True)

                # Check if STARTDATE is before ENDDATE
                for index, row in df.iterrows():
                    if row['STARTDATE'] >= row['ENDDATE']:
                        flash(f"Error: The STARTDATE in row {index + 2} must be before the ENDDATE.")
                        return redirect(url_for('indexTbCommissionLookup'))

                conn = get_db_connection()
                cursor = conn.cursor()

                # Prepare data for insertion
                new_entries = []
                for index, row in df.iterrows():
                    # Check for duplicates
                    cursor.execute("""
                        SELECT COUNT(*) FROM Z_TB_COMMISSION_LOOKUP
                        WHERE RETAILERID = ? AND PRODUCTID = ? AND STARTDATE = ? AND ENDDATE = ?
                    """, (row['RETAILERID'], row['PRODUCTID'], row['STARTDATE'], row['ENDDATE']))
                    exists = cursor.fetchone()[0]

                    if exists == 0:
                        new_entries.append((
                            row['RETAILERID'], row['PRODUCTID'], row['STARTDATE'], row['ENDDATE'], row['COMMISSIONVALUE']
                        ))

                # Batch insert new records
                if new_entries:
                    cursor.executemany("""
                        INSERT INTO Z_TB_COMMISSION_LOOKUP (RETAILERID, PRODUCTID, STARTDATE, ENDDATE, COMMISSIONVALUE)
                        VALUES (?, ?, ?, ?, ?)
                    """, new_entries)
                    conn.commit()
                    flash('Excel file data successfully processed and added to the database.')
                else:
                    flash('No new entries were added as all records already exist in the database.')

                cursor.close()
                conn.close()

            except Exception as e:
                flash(f'Error processing the file: {str(e)}')
        else:
            flash('Please upload a valid .xlsx file.')

        return redirect(url_for('indexTbCommissionLookup'))


    search_query = request.args.get('search')
    search_field = request.args.get('field', 'RETAILERID').upper()
    page = request.args.get('page', 1, type=int)
    per_page = 8

    conn = get_db_connection()
    cursor = conn.cursor()

    if search_query:
        query = f"SELECT COUNT(*) FROM Z_TB_COMMISSION_LOOKUP WHERE {search_field} LIKE ?"
        cursor.execute(query, (f"%{search_query}%",))
        total = cursor.fetchone()[0]

        query = f"""
        SELECT CAMPAIGNID, RETAILERID, PRODUCTID, STARTDATE, ENDDATE, COMMISSIONVALUE
        FROM Z_TB_COMMISSION_LOOKUP
        WHERE {search_field} LIKE ?
        ORDER BY CAMPAIGNID DESC
        LIMIT ? OFFSET ?
        """
        cursor.execute(query, (f"%{search_query}%", per_page, (page - 1) * per_page))
    else:
        query = "SELECT COUNT(*) FROM Z_TB_COMMISSION_LOOKUP"
        cursor.execute(query)
        total = cursor.fetchone()[0]

        query = """
        SELECT CAMPAIGNID, RETAILERID, PRODUCTID, STARTDATE, ENDDATE, COMMISSIONVALUE
        FROM Z_TB_COMMISSION_LOOKUP
        ORDER BY CAMPAIGNID DESC
        LIMIT ? OFFSET ?
        """
        cursor.execute(query, (per_page, (page - 1) * per_page))

    data = cursor.fetchall()
    cursor.close()
    conn.close()

    total_pages = math.ceil(total / per_page)

    if 'download' in request.args:
        return export_tb_commission_lookup_to_excel(search_query, search_field)

    return render_template('indexTbCommissionLookup.html', data=data, page=page, total_pages=total_pages, search_query=search_query, search_field=search_field)


def export_tb_commission_lookup_to_excel(search_query=None, search_field=None):
    conn = get_db_connection()
    cursor = conn.cursor()

    batch_size = 10000
    offset = 0
    all_rows = []

    while True:
        if search_query:
            query = f"""
            SELECT RETAILERID, PRODUCTID, STARTDATE, ENDDATE, COMMISSIONVALUE
            FROM Z_TB_COMMISSION_LOOKUP
            WHERE {search_field} LIKE ?
            ORDER BY CAMPAIGNID DESC
            LIMIT ? OFFSET ?
            """
            params = (f"%{search_query}%", batch_size, offset)
        else:
            query = """
            SELECT RETAILERID, PRODUCTID, STARTDATE, ENDDATE, COMMISSIONVALUE
            FROM Z_TB_COMMISSION_LOOKUP
            ORDER BY CAMPAIGNID DESC
            LIMIT ? OFFSET ?
            """
            params = (batch_size, offset)

        cursor.execute(query, params)
        rows = cursor.fetchall()

        if not rows:
            break

        all_rows.extend(rows)
        offset += batch_size

    conn.close()

    df = pd.DataFrame(all_rows, columns=[
        'RetailerID', 'ProductID', 'StartDate', 'EndDate', 'CommissionValue'
    ])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='TbCommissionLookup')

    output.seek(0)

    return send_file(output, as_attachment=True, download_name='TbCommissionLookupData.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/addTbCommissionLookupEntry', methods=['GET', 'POST'])
def addTbCommissionLookupEntry():
    if request.method == 'POST':
        retailerid = request.form['retailerid']
        productid = request.form['productid']
        startdate = request.form['startdate']
        enddate = request.form['enddate']
        commissionvalue = request.form['commissionvalue']

        # Validate date format
        try:
            startdate = datetime.strptime(startdate, '%Y-%m-%d')
            enddate = datetime.strptime(enddate, '%Y-%m-%d')
        except ValueError:
            flash("Invalid date format. Please use YYYY-MM-DD.")
            return redirect(url_for('addTbCommissionLookupEntry'))

        conn = get_db_connection()
        cursor = conn.cursor()

        try:
            # Insert new entry
            cursor.execute("""
                INSERT INTO Z_TB_COMMISSION_LOOKUP (RETAILERID, PRODUCTID, STARTDATE, ENDDATE, COMMISSIONVALUE)
                VALUES (?, ?, ?, ?, ?)
            """, (retailerid, productid, startdate, enddate, commissionvalue))

            conn.commit()
            flash('New TbCommissionLookup entry added successfully.')
        except Exception as e:
            flash(f'An error occurred: {str(e)}')
        finally:
            cursor.close()
            conn.close()

        return redirect(url_for('indexTbCommissionLookup'))

    return render_template('addTbCommissionLookupEntry.html')


@app.route('/editTbCommissionLookup/<int:campaignid>', methods=['GET', 'POST'])
def editTbCommissionLookup(campaignid):
    conn = get_db_connection()
    cursor = conn.cursor()

    if request.method == 'POST':
        new_retailerid = request.form['retailerid']
        new_productid = request.form['productid']
        new_startdate = request.form['startdate']
        new_enddate = request.form['enddate']
        new_commissionvalue = request.form['commissionvalue']

        # Validate date format
        try:
            new_startdate = datetime.strptime(new_startdate, '%Y-%m-%d')
            new_enddate = datetime.strptime(new_enddate, '%Y-%m-%d')
        except ValueError:
            flash("Invalid date format. Please use YYYY-MM-DD.")
            return redirect(url_for('editTbCommissionLookup', campaignid=campaignid))

        cursor.execute("""
            UPDATE Z_TB_COMMISSION_LOOKUP
            SET RETAILERID = ?, PRODUCTID = ?, STARTDATE = ?, ENDDATE = ?, COMMISSIONVALUE = ?
            WHERE CAMPAIGNID = ?
        """, (new_retailerid, new_productid, new_startdate, new_enddate, new_commissionvalue, campaignid))
        conn.commit()

        flash('Entry successfully updated.')
        return redirect(url_for('indexTbCommissionLookup'))

    cursor.execute("""
    SELECT CAMPAIGNID, RETAILERID, PRODUCTID, STARTDATE, ENDDATE, COMMISSIONVALUE
    FROM Z_TB_COMMISSION_LOOKUP
    WHERE CAMPAIGNID = ?
    """, (campaignid,))
    entry = cursor.fetchone()

    if entry:
        # Format the dates to YYYY-MM-DD format
        entry = list(entry)
        entry[3] = entry[3].strftime('%Y-%m-%d')
        entry[4] = entry[4].strftime('%Y-%m-%d')

    cursor.close()
    conn.close()

    if not entry:
        flash('Entry not found.')
        return redirect(url_for('indexTbCommissionLookup'))

    return render_template('editTbCommissionLookup.html', entry=entry)


@app.route('/deleteTbCommissionLookup/<int:campaignid>', methods=['POST'])
def deleteTbCommissionLookup(campaignid):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM Z_TB_COMMISSION_LOOKUP WHERE CAMPAIGNID = ?", (campaignid,))
    conn.commit()
    cursor.close()
    conn.close()
    flash('Entry successfully deleted.')
    return redirect(url_for('indexTbCommissionLookup'))


@app.route('/viewTbCommissionLookup/<int:campaignid>', methods=['GET'])
def viewTbCommissionLookup(campaignid):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Fetch the commission lookup's data
    cursor.execute("""
        SELECT CAMPAIGNID, RETAILERID, PRODUCTID, STARTDATE, ENDDATE, COMMISSIONVALUE
        FROM Z_TB_COMMISSION_LOOKUP
        WHERE CAMPAIGNID = ?
    """, (campaignid,))
    commission_data = cursor.fetchone()
    cursor.close()
    conn.close()

    if commission_data is None:
        flash('Commission lookup not found.', 'error')
        return redirect(url_for('indexTbCommissionLookup'))

    # Format the dates to YYYY-MM-DD format for display
    commission_data = list(commission_data)
    commission_data[3] = commission_data[3].strftime('%Y-%m-%d')
    commission_data[4] = commission_data[4].strftime('%Y-%m-%d')

    return render_template('viewTbCommissionLookup.html', commission=commission_data)


















if __name__ == '__main__':
   
  app.run(debug=True, port=5001)
